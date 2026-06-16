#!/usr/bin/env python3
"""
PaceMap Daily Pipeline
Runs automatically via GitHub Actions at 6:30pm UK time.
Fetches racecards, runs ANCHOR engine, generates Claude narratives, publishes JSON to GitHub.

Required environment variables (set as GitHub Secrets):
  ANTHROPIC_API_KEY
  GITHUB_TOKEN
  RACING_API_USER
  RACING_API_PASS
  GMAIL_APP_PASSWORD   (for tweet email -- create at myaccount.google.com/apppasswords)
  GMAIL_FROM           (your Gmail address, e.g. ppcjobber@gmail.com)
"""

import os
import sys

# Validate required env vars
for var in ['ANTHROPIC_API_KEY', 'GITHUB_TOKEN', 'RACING_API_USER', 'RACING_API_PASS']:
    if not os.environ.get(var):
        print(f'ERROR: {var} not set')
        sys.exit(1)

print('✅ Environment variables validated')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 2 — ANCHOR Engine: Core Setup + Comment Parser
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

EXCEL_FILE       = "/content/cheltenham_friday.xlsx"
RACE_SHEET       = "320 Cheltenham"
RACE_NAME        = "Albert Bartlett Novices' Hurdle"
RACE_DATE        = "13 March 2026"

TODAY_GOING      = "GOOD_TO_SOFT"
TODAY_COURSE     = "CHE"
TODAY_DISTANCE_F = 22.25

STEPPING_UP_TRIP_HORSES  = []
STEPPING_UP_CLASS_HORSES = []
MANUAL_OVERRIDES         = {}

ODDS_MAP = {}

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.run(["pip","install","openpyxl","-q"])
    import openpyxl

from datetime import date, datetime
import re

FC_CLOTH   = 0
FC_HORSE   = 1
FC_AGE     = 2
FC_WGT     = 3
FC_OR      = 4
FC_JOCKEY  = 5
FC_TRAINER = 6
FC_DATE    = 7
FC_CONDS   = 8
FC_WGT2    = 9
FC_POS     = 10
FC_MARGIN  = 11
FC_SP      = 12
FC_JOC2    = 13
FC_TS      = 14
FC_RPR     = 15
FC_IRC     = 16
FC_RPC     = 17

_GOING_MAP = {
    "firm":         "FIRM",
    "gd/fm":"GOOD_TO_FIRM", "gd/f":"GOOD_TO_FIRM", "gf":"GOOD_TO_FIRM",
    "good to firm": "GOOD_TO_FIRM",
    "gd":           "GOOD",   "good":"GOOD",
    "gd/y":         "GOOD_TO_SOFT", "gd/yl":"GOOD_TO_SOFT",
    "good to soft": "GOOD_TO_SOFT", "good/soft":"GOOD_TO_SOFT",
    "yielding":     "GOOD_TO_SOFT", "gs":"GOOD_TO_SOFT", "y":"GOOD_TO_SOFT",
    "soft":         "SOFT",   "sft":"SOFT",
    "sft/hy":       "SOFT",   "soft/heavy":"SOFT",
    "hy":           "HEAVY",  "hvy":"HEAVY", "heavy":"HEAVY",
    "vsft":         "HEAVY",  "v soft":"HEAVY",
    "st":           "GOOD_TO_SOFT",
}

def norm_going(raw):
    if not raw: return "UNKNOWN"
    s = str(raw).strip().lower()
    if s in _GOING_MAP: return _GOING_MAP[s]
    for k in sorted(_GOING_MAP, key=len, reverse=True):
        if k in s: return _GOING_MAP[k]
    return "UNKNOWN"

_GRADE_MAP = [
    (r"nvhg1|nvg1|nov.*g1|nhg1|3yh",          "G1"),
    (r"nvhg2|nvg2|nov.*g2|nhg2|g2",            "G2"),
    (r"nvhg3|nvg3|nov.*g3|nhg3|g3",            "G3"),
    (r"cls1|c1|listed|lst",                     "CLS1"),
    (r"cls2|c2",                                "CLS2"),
    (r"cls3|c3",                                "CLS3"),
    (r"nv[hc]|novh|novice|nov\b|\bnv\b",        "NOV"),
    (r"hcap|hcp|handicap|\bhc\b|\bhchl\b",      "HCP"),
    (r"mdh|mdn|maiden",                         "MDN"),
    (r"nhf|bumper|nf\b|\bnhf\b|5md|4mdn",       "NHF"),
    (r"p2p|ptp|point|hunt",                     "P2P"),
]

def extract_grade(conds):
    if not conds: return "UNK"
    s = str(conds).lower()
    for pat, code in _GRADE_MAP:
        if re.search(pat, s): return code
    return "UNK"

_COURSE_MAP = {
    "che":"CHE","chel":"CHE","asc":"ASC","leo":"LEO","leop":"LEO",
    "pun":"PUN","punc":"PUN","nav":"NAV","nava":"NAV","naa":"NAA","naas":"NAA",
    "gal":"GAL","galw":"GAL","cor":"COR","cork":"COR","tip":"TIP","tipp":"TIP",
    "lim":"LIM","lime":"LIM","lis":"LIS","list":"LIS","kil":"KIL","kild":"KIL",
    "dow":"DOW","down":"DOW","gow":"GOW","gowr":"GOW","ros":"ROS","sli":"SLI",
    "bel":"BEL","bell":"BEL","bal":"BAL","bali":"BAL","clo":"CLO","clon":"CLO",
    "mar":"MAR","thu":"THU","mus":"MUS","muss":"MUS","ayr":"AYR","san":"SAN",
    "sand":"SAN","new":"NEW","newb":"NEW","hay":"HAY","hayd":"HAY","exe":"EXE",
    "exet":"EXE","tau":"TAU","taun":"TAU","wet":"WET","weth":"WET","don":"DON",
    "donc":"DON","yor":"YOR","york":"YOR","utt":"UTT","utto":"UTT","wor":"WOR",
    "worc":"WOR","lud":"LUD","ludl":"LUD","ban":"BAN","bang":"BAN","hex":"HEX",
    "hext":"HEX","sed":"SED","sedg":"SED","fon":"FON","font":"FON","win":"WIN",
    "winc":"WIN","war":"WAR","warw":"WAR","kem":"KEM","kemp":"KEM","lin":"LIN",
    "ling":"LIN","per":"PER","pert":"PER","aut":"AUT","dun":"DUN","dro":"DRO",
    "chp":"CHP","crl":"CRL","wrd":"WRD","wdr":"WDR","chl":"CHL","lrg":"LRG",
    "nwc":"NWC","mkt":"MKT",
}

def extract_course(conds):
    if not conds: return None
    token = str(conds).strip().split()[0].lower()
    for length in range(min(4, len(token)), 1, -1):
        prefix = token[:length]
        if prefix in _COURSE_MAP:
            return _COURSE_MAP[prefix]
    return token.upper()[:3]

def extract_distance_f(conds):
    if not conds: return None
    parts = str(conds).strip().split()
    if len(parts) < 2: return None
    raw = parts[1].replace("\u00bd",".5").replace("\u00bc",".25").replace("\u00be",".75")
    try:
        return float(raw)
    except ValueError:
        return None

_TEXT_MARGINS = {
    "dht":0.0,"dead heat":0.0,"nse":0.05,"nose":0.05,"sht-hd":0.05,"sh-hd":0.05,
    "hd":0.1,"head":0.1,"nk":0.25,"neck":0.25,"1/2l":0.5,
}
_NON_COMPLETION = {"pu","f","ur","u","bd","co","rr","ref","ro","nf","fell",
                   "brought down","unseated","pulled up","refused"}

def parse_margin_winner(raw):
    if not raw: return None, None
    s = str(raw).strip()
    for txt, val in _TEXT_MARGINS.items():
        if s.lower().startswith(txt):
            rest = s[len(txt):].strip()
            rival_match = re.match(r'(.+?)\s+\d{1,2}-\d{2}', rest)
            rival = rival_match.group(1).strip() if rival_match else rest
            return val, rival
    m = re.match(r'([\u00bd\u00bc\u00be]|\d+[\u00bd\u00bc\u00be]?(?:\.\d+)?)\s*[Ll]?\s+(.+)', s)
    if m:
        raw_m = m.group(1).replace("\u00bd",".5").replace("\u00bc",".25").replace("\u00be",".75")
        try:
            margin = float(raw_m)
        except ValueError:
            margin = None
        rest = m.group(2).strip()
        rival_match = re.match(r'(.+?)\s+\d{1,2}-\d{2}', rest)
        rival = rival_match.group(1).strip() if rival_match else rest
        return margin, rival
    return None, None

def parse_position(raw):
    if not raw: return None, None
    s = str(raw).strip().lower()
    if s in _NON_COMPLETION:
        return None, s.upper()
    s2 = re.sub(r'(st|nd|rd|th)$', '', s)
    try:
        return int(s2), None
    except ValueError:
        return None, s.upper()

def parse_date(raw):
    if not raw: return date.today()
    if isinstance(raw, datetime): return raw.date()
    if isinstance(raw, date):     return raw
    s = str(raw).strip()
    for fmt in ("%d%b%y","%d%b%Y","%d/%m/%Y","%d/%m/%y","%d-%m-%Y","%d %b %Y","%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return date.today()

RPR_PROXY_DISCOUNT = 0.97

def _build_signal_from_irc(irc_text, grade_weight):
    if not irc_text:
        return parse_comment("", grade_weight)
    return parse_comment(str(irc_text), grade_weight)

def _build_anchors(position, margin, rpr, horse_or, rival_name, field_or_map):
    if rival_name and field_or_map:
        for name, or_val in field_or_map.items():
            if (name.lower() in rival_name.lower() or rival_name.lower() in name.lower()):
                if position == 1:
                    return or_val, None
                else:
                    return None, or_val
    rpr_float = None
    try:
        rpr_float = float(str(rpr)) if rpr and str(rpr).strip() not in ("-","") else None
    except (ValueError, TypeError):
        pass
    if rpr_float:
        proxy = int(rpr_float * RPR_PROXY_DISCOUNT)
        return None, proxy
    try:
        or_int = int(float(str(horse_or))) if horse_or and str(horse_or).strip() not in ("-","") else None
    except (ValueError, TypeError):
        or_int = None
    return None, or_int

def load_race(excel_file, sheet_name, today_going, today_course, today_distance_f,
              stepping_up_trip=None, stepping_up_class=None):
    stepping_up_trip  = stepping_up_trip  or []
    stepping_up_class = stepping_up_class or []
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        match = next((s for s in wb.sheetnames if sheet_name.lower() in s.lower()), None)
        if match:
            ws = wb[match]
            print(f"  Sheet '{sheet_name}' not found — using '{match}'")
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    all_rows = list(ws.iter_rows(values_only=True))
    data_start = 2
    for i, row in enumerate(all_rows[:5]):
        if row[0] and str(row[0]).strip().lower() in ("no", "no.", "#"):
            data_start = i + 1
            break
    rows = all_rows[data_start:]
    numbered = [(i, row) for i, row in enumerate(rows)
                if row[FC_CLOTH] and str(row[FC_CLOTH]).strip().isdigit()]
    runner_row_indices = set()
    prev_date = None
    for i, row in numbered:
        this_date = row[FC_DATE]
        if this_date != prev_date:
            runner_row_indices.add(i)
            prev_date = this_date
    field_or_map = {}
    for i, row in numbered:
        if i in runner_row_indices:
            name = str(row[FC_HORSE]).strip() if row[FC_HORSE] else None
            or_v = row[FC_OR]
            if name and or_v and str(or_v).strip() not in ("-",""):
                try:
                    field_or_map[name] = int(float(str(or_v)))
                except (ValueError, TypeError):
                    pass
    horses = []; load_log = []; proxied_or_map = {}
    current_meta = None; form_rows = []

    def finalise_horse(meta, frows):
        name, cloth, or_val, jockey, trainer = meta
        or_int = None; proxy_note = None
        try:
            or_int = int(float(str(or_val))) if or_val and str(or_val).strip() not in ("-","") else None
        except (ValueError, TypeError):
            pass
        if not or_int:
            rprs = []
            for fr in frows:
                r = fr[FC_RPR]
                if r and str(r).strip() not in ("-",""):
                    try: rprs.append(int(float(str(r))))
                    except: pass
            if rprs:
                best_rpr = max(rprs)
                or_int   = int(best_rpr * RPR_PROXY_DISCOUNT)
                proxy_note = f"best RPR {best_rpr} \u00d7 {RPR_PROXY_DISCOUNT} = {or_int}"
            else:
                or_int = 120; proxy_note = "no RPR available — fallback OR 120"
        runs = []
        for fr in frows:
            run_date = parse_date(fr[FC_DATE]); conds = fr[FC_CONDS]
            pos_raw = fr[FC_POS]; margin_raw = fr[FC_MARGIN]
            rpr = fr[FC_RPR]; irc = fr[FC_IRC]
            grade = extract_grade(conds); course = extract_course(conds)
            dist_f = extract_distance_f(conds)
            going = norm_going(str(conds).split()[2] if conds and len(str(conds).split()) > 2 else "")
            position, comp_code = parse_position(pos_raw)
            margin, rival = parse_margin_winner(margin_raw)
            gw = GRADE_WEIGHTS.get(grade, GRADE_WEIGHTS.get("UNK", 0.55))
            second_or, race_or_anchor = _build_anchors(position, margin, rpr, or_int, rival, field_or_map)
            signal = _build_signal_from_irc(irc, gw)
            fr_obj = FormRun(run_date=run_date, position=position, completion_code=comp_code,
                             margin_lengths=margin, second_or=second_or, race_or_anchor=race_or_anchor,
                             grade_code=grade, signal=signal, chain_depth=0, chain_source_or=None)
            fr_obj.going_category = going; fr_obj.course_code = course; fr_obj.distance_f = dist_f
            runs.append(fr_obj)
        h = Horse(name=name, official_rating=or_int, cloth_number=cloth, runs=runs)
        h._proxy_note = proxy_note
        return h, proxy_note

    for i, row in enumerate(rows):
        if i in runner_row_indices:
            if current_meta:
                h, pnote = finalise_horse(current_meta, form_rows)
                horses.append(h)
                pflag = " [OR proxied*]" if pnote else ""
                load_log.append(f"  \u2713 {h.name:<28} OR={h.official_rating:<4}  {len(h.runs)} run(s){pflag}")
                if pnote: proxied_or_map[h.name] = pnote
            name = str(row[FC_HORSE]).strip() if row[FC_HORSE] else "Unknown"
            cloth = int(str(row[FC_CLOTH]).strip())
            or_val = row[FC_OR]; jockey = row[FC_JOCKEY]; trainer = row[FC_TRAINER]
            current_meta = (name, cloth, or_val, jockey, trainer)
            form_rows = [row] if row[FC_DATE] else []
        elif current_meta and not row[FC_CLOTH] and row[FC_DATE]:
            form_rows.append(row)
        elif row[FC_CLOTH] and str(row[FC_CLOTH]).strip().isdigit():
            pass

    if current_meta:
        h, pnote = finalise_horse(current_meta, form_rows)
        horses.append(h)
        pflag = " [OR proxied*]" if pnote else ""
        load_log.append(f"  \u2713 {h.name:<28} OR={h.official_rating:<4}  {len(h.runs)} run(s){pflag}")
        if pnote: proxied_or_map[h.name] = pnote

    step_trip_map  = {h: True for h in stepping_up_trip}
    step_class_map = {h: True for h in stepping_up_class}
    return horses, step_trip_map, step_class_map, load_log, proxied_or_map

def run_engine(excel_file, sheet_name, race_name, race_date, today_going, today_course,
               today_distance_f, odds_map, stepping_up_trip=None, stepping_up_class=None,
               override_map=None, show_dampening_trail=True, race_grade='DEFAULT'):
    from datetime import date as _d
    print(f"\n{'═'*65}")
    print(f"  ANCHOR — {race_name}")
    print(f"  {race_date}  |  {today_course}  |  {today_distance_f}f  |  {today_going}")
    print(f"{'═'*65}")
    horses, step_trip, step_class, log, proxied_or_map = load_race(
        excel_file, sheet_name, today_going, today_course, today_distance_f,
        stepping_up_trip, stepping_up_class)
    for line in log: print(line)
    print(f"\n  {len(horses)} horses loaded.")
    today = _d.today()
    avg_or, sd = field_stats(horses)
    print(f"  Field avg OR: {avg_or:.1f}  |  SD: {sd:.1f}  |  Ceiling: {avg_or + 2*sd:.1f}\n")
    rr_map = {h.name: calculate_private_rating(h, avg_or, sd, today) for h in horses}
    pref_map = {h.name: calculate_preferences(h, rr_map[h.name].run_ratings,
                today_going, today_course, today_distance_f) for h in horses}
    pace = classify_pace(horses, {h.name: rr_map[h.name].run_ratings for h in horses})
    damp_map = dampen_field(horses=horses, rating_results_map=rr_map, pref_results_map=pref_map,
                            pace_result=pace, field_avg_or=avg_or,
                            step_up_trip_map=step_trip, step_up_class_map=step_class)
    ev_summary = calculate_ev(dampening_results=damp_map, odds_map=odds_map, race_name=race_name,
                              field_avg_or=avg_or, pace_result=pace, override_map=override_map or {},
                              proxied_or_map=proxied_or_map, race_grade=race_grade,
                              distance_f=today_distance_f)
    trail = damp_map if show_dampening_trail else None
    qcard = quick_card(ev_summary)
    freport = full_report(ev_summary, dampening_map=trail, race_date=race_date)
    print(qcard); print(); print(freport)
    return ev_summary, freport, qcard

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 3 — ANCHOR Engine: Comment Parser
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from dataclasses import dataclass, field
from enum import Enum
import re

class RunningStyle(Enum):
    FRONT_RUNNER  = "FRONT_RUNNER"
    PROMINENT     = "PROMINENT"
    MIDFIELD      = "MIDFIELD"
    HOLD_UP       = "HOLD_UP"
    UNKNOWN       = "UNKNOWN"

class TripSignal(Enum):
    WANTS_FURTHER  = "WANTS_FURTHER"
    AT_OPTIMUM     = "AT_OPTIMUM"
    TRIP_TOO_FAR   = "TRIP_TOO_FAR"
    TRIP_TOO_SHORT = "TRIP_TOO_SHORT"
    UNKNOWN        = "UNKNOWN"

class GoingSignal(Enum):
    SUITED   = "SUITED"
    NEUTRAL  = "NEUTRAL"
    UNSUITED = "UNSUITED"
    UNKNOWN  = "UNKNOWN"

class JumpingSignal(Enum):
    FLUENT      = "FLUENT"
    NEUTRAL     = "NEUTRAL"
    DIRECTIONAL = "DIRECTIONAL"
    POOR        = "POOR"
    UNKNOWN     = "UNKNOWN"

class PerformanceSignal(Enum):
    IMPRESSIVE = "IMPRESSIVE"
    POSITIVE   = "POSITIVE"
    NEUTRAL    = "NEUTRAL"
    BELOW_PAR  = "BELOW_PAR"
    EXCUSED    = "EXCUSED"

class DirectionalFlag(Enum):
    LEFT  = "LEFT"
    RIGHT = "RIGHT"
    NONE  = "NONE"

@dataclass
class SignalObject:
    running_style:       RunningStyle      = RunningStyle.UNKNOWN
    trip_signal:         TripSignal        = TripSignal.UNKNOWN
    going_signal:        GoingSignal       = GoingSignal.UNKNOWN
    jumping_signal:      JumpingSignal     = JumpingSignal.UNKNOWN
    performance_signal:  PerformanceSignal = PerformanceSignal.NEUTRAL
    confidence_modifier: float             = 1.00
    excuse_flag:         bool              = False
    directional_flag:    DirectionalFlag   = DirectionalFlag.NONE
    matched_phrases:     list              = field(default_factory=list)

RUNNING_STYLE_PHRASES = [
    (r"made (all|virtually all|every yard)",          RunningStyle.FRONT_RUNNER, 1.00),
    (r"\bled (from|throughout|early|the way)",        RunningStyle.FRONT_RUNNER, 1.00),
    (r"set (the |a )?(strong |fierce |good )?pace",   RunningStyle.FRONT_RUNNER, 1.00),
    (r"went off (in )?front",                         RunningStyle.FRONT_RUNNER, 1.00),
    (r"disputed (the )?lead",                         RunningStyle.FRONT_RUNNER, 1.00),
    (r"tracked (the |a )?(leader|pace|winner)",       RunningStyle.PROMINENT,    1.00),
    (r"chased (the |a )?(leader|pace|winner)",        RunningStyle.PROMINENT,    1.00),
    (r"pressed (the |a )?(leader|pace)",              RunningStyle.PROMINENT,    1.00),
    (r"prominent(ly)?",                               RunningStyle.PROMINENT,    1.00),
    (r"close (up |to )(the |a )?(leader|pace)",       RunningStyle.PROMINENT,    1.00),
    (r"mid.?division",                                RunningStyle.MIDFIELD,     1.00),
    (r"in (the )?middle",                             RunningStyle.MIDFIELD,     1.00),
    (r"in touch",                                     RunningStyle.MIDFIELD,     1.00),
    (r"\bhandy\b",                                    RunningStyle.MIDFIELD,     1.00),
    (r"held (a )?prominent position",                 RunningStyle.MIDFIELD,     1.00),
    (r"held up",                                      RunningStyle.HOLD_UP,      1.00),
    (r"towards (the )?rear",                          RunningStyle.HOLD_UP,      1.00),
    (r"dropped (out|away|to (the )?rear)",            RunningStyle.HOLD_UP,      1.00),
    (r"in (the )?rear",                               RunningStyle.HOLD_UP,      1.00),
    (r"last (of all|pair|three)",                     RunningStyle.HOLD_UP,      1.00),
    (r"settled (in )?last",                           RunningStyle.HOLD_UP,      1.00),
]

TRIP_PHRASES = [
    (r"ran on strongly",                              TripSignal.WANTS_FURTHER,  1.10),
    (r"running on (at|towards) (the )?finish",        TripSignal.WANTS_FURTHER,  1.10),
    (r"stayed on well",                               TripSignal.WANTS_FURTHER,  1.10),
    (r"kept (on|finding)",                            TripSignal.WANTS_FURTHER,  1.10),
    (r"found plenty",                                 TripSignal.WANTS_FURTHER,  1.10),
    (r"staying on (at|at the) finish",                TripSignal.WANTS_FURTHER,  1.08),
    (r"stayed on (one pace)?",                        TripSignal.WANTS_FURTHER,  1.05),
    (r"never able to land a blow",                    TripSignal.WANTS_FURTHER,  0.90),
    (r"weakened (quickly|rapidly|from|after|before)", TripSignal.TRIP_TOO_FAR,   1.10),
    (r"flattened out",                                TripSignal.TRIP_TOO_FAR,   1.10),
    (r"stopped (quickly|badly|to nothing)",           TripSignal.TRIP_TOO_FAR,   1.10),
    (r"ran out of (petrol|steam|energy)",             TripSignal.TRIP_TOO_FAR,   1.10),
    (r"outpaced (after|from|before)",                 TripSignal.TRIP_TOO_FAR,   1.05),
    (r"found little (extra|more)",                    TripSignal.TRIP_TOO_FAR,   1.05),
    (r"nothing (left|more) to give",                  TripSignal.TRIP_TOO_FAR,   1.05),
    (r"tired (badly|quickly|from)",                   TripSignal.TRIP_TOO_FAR,   1.05),
    (r"could not quicken",                            TripSignal.TRIP_TOO_SHORT, 1.05),
    (r"unable to quicken",                            TripSignal.TRIP_TOO_SHORT, 1.05),
    (r"never got competitive",                        TripSignal.TRIP_TOO_SHORT, 0.90),
]

GOING_PHRASES = [
    (r"relished (the )?(conditions|ground|going)",         GoingSignal.SUITED,   1.10),
    (r"suited (by|to) (the )?(going|ground|conditions)",   GoingSignal.SUITED,   1.10),
    (r"in (his|her|its) element",                          GoingSignal.SUITED,   1.10),
    (r"travelled (well|strongly)",                         GoingSignal.SUITED,   1.05),
    (r"handled (the )?(conditions|ground|going)",          GoingSignal.SUITED,   1.05),
    (r"unsuited (by|to) (the )?(going|ground|conditions)", GoingSignal.UNSUITED, 1.10),
    (r"never travelling (on|well)",                        GoingSignal.UNSUITED, 1.10),
    (r"didn.t handle",                                     GoingSignal.UNSUITED, 1.10),
    (r"lost (footing|his|her) footing",                    GoingSignal.UNSUITED, 1.10),
    (r"slipped (up|badly|on)",                             GoingSignal.UNSUITED, 1.10),
    (r"never at home (on|in)",                             GoingSignal.UNSUITED, 1.10),
]

JUMPING_PHRASES = [
    (r"jumped (well|boldly|fluently|cleanly|great)",          JumpingSignal.FLUENT,       1.10),
    (r"fluent (jumper|jumping|over)",                         JumpingSignal.FLUENT,       1.10),
    (r"stood off well",                                       JumpingSignal.FLUENT,       1.10),
    (r"bold (jumper|jumping)",                                JumpingSignal.FLUENT,       1.10),
    (r"jumped (consistently |persistently )?right",           JumpingSignal.DIRECTIONAL,  1.00),
    (r"jumped (consistently |persistently )?left",            JumpingSignal.DIRECTIONAL,  1.00),
    (r"hung (right|left) (at|over|after)",                    JumpingSignal.DIRECTIONAL,  1.00),
    (r"nearly fell",                                          JumpingSignal.POOR,         0.75),
    (r"blundered (badly|at|seriously)",                       JumpingSignal.POOR,         0.75),
    (r"bad (mistake|error|blunder)",                          JumpingSignal.POOR,         0.80),
    (r"(made a |)mistake (at|before|after)",                  JumpingSignal.POOR,         0.90),
    (r"hit (the top of |)(the )?(fence|hurdle|last|final)",   JumpingSignal.POOR,         0.90),
    (r"not fluent",                                           JumpingSignal.POOR,         0.90),
]

PERFORMANCE_PHRASES = [
    (r"(on the |on )?bridle",                                 PerformanceSignal.IMPRESSIVE, 1.25),
    (r"cruising",                                             PerformanceSignal.IMPRESSIVE, 1.25),
    (r"cantered (home|clear|away)",                           PerformanceSignal.IMPRESSIVE, 1.25),
    (r"eased down",                                           PerformanceSignal.IMPRESSIVE, 1.20),
    (r"impressive(ly)?",                                      PerformanceSignal.IMPRESSIVE, 1.20),
    (r"idling (late|in front|close home)",                    PerformanceSignal.IMPRESSIVE, 1.20),
    (r"very (easily|comfortably)",                            PerformanceSignal.IMPRESSIVE, 1.20),
    (r"going well",                                           PerformanceSignal.POSITIVE,   1.15),
    (r"always holding",                                       PerformanceSignal.POSITIVE,   1.10),
    (r"always doing enough",                                  PerformanceSignal.POSITIVE,   1.10),
    (r"comfortable(ly)?",                                     PerformanceSignal.POSITIVE,   1.10),
    (r"responded well (to pressure|to the|under)",            PerformanceSignal.POSITIVE,   1.10),
    (r"kept finding",                                         PerformanceSignal.POSITIVE,   1.10),
    (r"(easily|readily|without (too much|much) (effort|trouble))", PerformanceSignal.POSITIVE, 1.10),
    (r"well held",                                            PerformanceSignal.BELOW_PAR,  0.85),
    (r"ran below (form|expectations|par)",                    PerformanceSignal.BELOW_PAR,  0.85),
    (r"disappointing(ly)?",                                   PerformanceSignal.BELOW_PAR,  0.85),
    (r"one.?paced",                                           PerformanceSignal.BELOW_PAR,  0.90),
    (r"never (really )?dangerous",                            PerformanceSignal.BELOW_PAR,  0.90),
    (r"no extra",                                             PerformanceSignal.BELOW_PAR,  0.90),
    (r"plugged on",                                           PerformanceSignal.BELOW_PAR,  0.92),
]

EXCUSE_PHRASES = [
    r"hampered", r"badly hampered", r"checked",
    r"knocked (sideways|over|into)", r"impeded",
    r"short of room", r"no room", r"squeezed out",
    r"brought down", r"carried out",
    r"lost ground at (the )?start", r"slowly away",
    r"missed (the )?break", r"broke (a )?blood vessel", r"bled",
]

def _match_first(text, phrases):
    for pattern, signal, modifier in phrases:
        if re.search(pattern, text, re.IGNORECASE):
            return signal, modifier, pattern
    return None, None, None

def _match_all(text, phrases):
    return [(signal, modifier, pattern)
            for pattern, signal, modifier in phrases
            if re.search(pattern, text, re.IGNORECASE)]

def _detect_directional(text):
    if re.search(r"jumped (consistently |persistently )?right|hung right", text, re.IGNORECASE):
        return DirectionalFlag.RIGHT
    if re.search(r"jumped (consistently |persistently )?left|hung left", text, re.IGNORECASE):
        return DirectionalFlag.LEFT
    return DirectionalFlag.NONE

def parse_comment(comment, grade_wt=1.0):
    if not comment or not comment.strip():
        return SignalObject()
    text = comment.strip(); sig = SignalObject(); matched = []
    excuse_triggers = [p for p in EXCUSE_PHRASES if re.search(p, text, re.IGNORECASE)]
    if excuse_triggers:
        sig.excuse_flag = True; sig.confidence_modifier = 0.60
        sig.performance_signal = PerformanceSignal.EXCUSED
        style, _, sp = _match_first(text, RUNNING_STYLE_PHRASES)
        if style: sig.running_style = style
        sig.matched_phrases = [f"EXCUSE: {p}" for p in excuse_triggers]
        return sig
    style, _, sp = _match_first(text, RUNNING_STYLE_PHRASES)
    if style: sig.running_style = style; matched.append(f"style: {sp}")
    trip_matches = _match_all(text, TRIP_PHRASES); trip_conf = 1.00
    if trip_matches:
        best = max(trip_matches, key=lambda x: abs(x[1] - 1.0))
        sig.trip_signal = best[0]; matched.append(f"trip: {best[2]}"); trip_conf = best[1]
    going, going_conf, gp = _match_first(text, GOING_PHRASES); going_conf = going_conf or 1.00
    if going: sig.going_signal = going; matched.append(f"going: {gp}")
    jump_matches = _match_all(text, JUMPING_PHRASES); jump_conf = 1.00
    if jump_matches:
        poor = [m for m in jump_matches if m[0] == JumpingSignal.POOR]
        direc = [m for m in jump_matches if m[0] == JumpingSignal.DIRECTIONAL]
        fluent = [m for m in jump_matches if m[0] == JumpingSignal.FLUENT]
        if poor:
            worst = min(poor, key=lambda x: x[1])
            sig.jumping_signal = JumpingSignal.POOR; matched.append(f"jump: {worst[2]}"); jump_conf = worst[1]
        elif direc:
            sig.jumping_signal = JumpingSignal.DIRECTIONAL
            sig.directional_flag = _detect_directional(text); matched.append(f"jump: {direc[0][2]}")
        else:
            best = max(fluent, key=lambda x: x[1])
            sig.jumping_signal = JumpingSignal.FLUENT; matched.append(f"jump: {best[2]}"); jump_conf = best[1]
    perf_matches = _match_all(text, PERFORMANCE_PHRASES); perf_conf = 1.00
    if perf_matches:
        priority = {PerformanceSignal.IMPRESSIVE:4, PerformanceSignal.POSITIVE:3,
                    PerformanceSignal.NEUTRAL:2, PerformanceSignal.BELOW_PAR:1}
        best = max(perf_matches, key=lambda x: priority.get(x[0], 0))
        sig.performance_signal = best[0]; matched.append(f"perf: {best[2]}"); perf_conf = best[1]
    combined = trip_conf * going_conf * jump_conf * perf_conf
    combined = max(0.60, min(1.25, combined))
    grade_ceiling = 1.0 + (grade_wt * 0.25)
    combined = min(combined, grade_ceiling)
    sig.confidence_modifier = round(combined, 3); sig.matched_phrases = matched
    return sig

def modal_running_style(signals, max_comments=3):
    from collections import Counter
    recent = signals[:max_comments]
    styles = [s.running_style for s in recent if s.running_style != RunningStyle.UNKNOWN]
    if not styles: return RunningStyle.UNKNOWN
    counts = Counter(styles); modal = counts.most_common(1)[0]
    if modal[1] == 1 and len(styles) > 1: return styles[0]
    return modal[0]

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 3 — Private Ratings Engine
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from dataclasses import dataclass, field
from datetime import date
from typing import Optional
import math

LB_PER_LENGTH   = 1.0
MAX_CHAIN_DEPTH = 3

RECENCY_TABLE = [
    (0,   1.00), (50,  1.00), (77,  0.85),
    (118, 0.70), (146, 0.58), (180, 0.46),
    (307, 0.25), (365, 0.15),
]

GRADE_WEIGHTS = {
    "G1": 1.00, "G2": 0.90, "G3": 0.80,
    "CLS1": 0.75, "CLS2": 0.65, "CLS3": 0.65,
    "NOV": 0.55, "MDN": 0.55,
    "NHF": 0.40, "P2P": 0.30, "UNK": 0.55,
}

CHAIN_DISCOUNTS = {0: 1.00, 1: 0.90, 2: 0.75, 3: 0.60}
NON_COMPLETION  = {"PU", "F", "U", "BD", "RR", "REF", "CO", "UR", "NF"}

@dataclass
class FormRun:
    run_date:          date
    position:          Optional[int]
    completion_code:   Optional[str]
    margin_lengths:    Optional[float]
    second_or:         Optional[int]
    race_or_anchor:    Optional[int]
    grade_code:        str   = "UNK"
    signal:            object = None
    chain_depth:       int   = 0
    chain_source_or:   Optional[int] = None

@dataclass
class Horse:
    name:            str
    official_rating: int
    cloth_number:    int   = 0
    runs:            list  = field(default_factory=list)

@dataclass
class RunRating:
    run:             FormRun
    private_rating:  float
    capped:          bool
    recency_weight:  float
    grade_weight:    float
    confidence_mod:  float
    chain_discount:  float
    run_weight:      float

@dataclass
class PrivateRatingResult:
    horse_name:      str
    official_rating: int
    run_ratings:     list
    weighted_avg:    float
    qualifying_runs: int
    avg_grade_weight:float
    max_chain_depth: int
    field_avg_or:    float
    field_ceiling:   float

    def summary(self):
        lines = [
            f"\n{'═'*55}",
            f"  {self.horse_name}  (OR {self.official_rating})",
            f"{'─'*55}",
            f"  Private Rating  : {self.weighted_avg:.1f}  "
            f"({'+'if self.weighted_avg >= self.official_rating else ''}"
            f"{self.weighted_avg - self.official_rating:.1f}lb vs OR)",
            f"  Qualifying runs : {self.qualifying_runs}",
            f"  Avg grade wt    : {self.avg_grade_weight:.2f}",
            f"  Max chain depth : {self.max_chain_depth}",
            f"  Field ceiling   : {self.field_ceiling:.1f}",
            f"{'─'*55}  Run breakdown:",
        ]
        for rr in self.run_ratings:
            cap = " [CAP]" if rr.capped else ""
            lines.append(
                f"    {rr.run.run_date}  Rtg {rr.private_rating:.1f}{cap}  "
                f"Wt {rr.run_weight:.3f}  "
                f"(rec {rr.recency_weight:.2f} x grd {rr.grade_weight:.2f} x "
                f"conf {rr.confidence_mod:.2f} x chn {rr.chain_discount:.2f})")
        lines.append(f"{'═'*55}")
        return "\n".join(lines)

def get_recency_weight(run_date, today=None):
    if today is None: today = date.today()
    days = max(0, (today - run_date).days)
    if days >= RECENCY_TABLE[-1][0]: return RECENCY_TABLE[-1][1]
    for i in range(len(RECENCY_TABLE) - 1):
        d0, w0 = RECENCY_TABLE[i]; d1, w1 = RECENCY_TABLE[i + 1]
        if d0 <= days <= d1:
            t = (days - d0) / (d1 - d0)
            return round(w0 + t * (w1 - w0), 4)
    return RECENCY_TABLE[0][1]

def get_grade_weight(grade_code):
    return GRADE_WEIGHTS.get(grade_code.upper(), GRADE_WEIGHTS["UNK"])

def get_chain_discount(depth):
    if depth > MAX_CHAIN_DEPTH: return 0.0
    return CHAIN_DISCOUNTS.get(depth, 0.0)

def is_non_completion(run):
    if run.completion_code and run.completion_code.upper() in NON_COMPLETION: return True
    if run.position is None and run.completion_code: return True
    return False

def raw_private_rating(run, horse_or):
    if is_non_completion(run): return None
    cd = get_chain_discount(run.chain_depth)
    if cd == 0.0: return None
    if run.position == 1 and run.margin_lengths is not None:
        anchor = run.second_or or run.race_or_anchor or run.chain_source_or
        if anchor is None: return None
        pr = float(anchor) + (run.margin_lengths * LB_PER_LENGTH)
        if run.chain_depth > 0: pr = (pr * cd) + (horse_or * (1.0 - cd))
        return pr
    if run.position and run.position > 1 and run.margin_lengths is not None:
        anchor = run.race_or_anchor or run.chain_source_or
        if anchor is None: return None
        pr = float(anchor) - (run.margin_lengths * LB_PER_LENGTH)
        if run.chain_depth > 0: pr = (pr * cd) + (horse_or * (1.0 - cd))
        return pr
    return None

def field_stats(horses):
    ors = [h.official_rating for h in horses if h.official_rating]
    if not ors: return 0.0, 0.0
    avg = sum(ors) / len(ors)
    sd  = math.sqrt(sum((x - avg) ** 2 for x in ors) / len(ors))
    return round(avg, 2), round(sd, 2)

def calculate_private_rating(horse, field_avg_or, field_sd, today=None):
    if today is None: today = date.today()
    ceiling = field_avg_or + (2.0 * field_sd)
    run_ratings = []; total_weight = 0.0; weighted_sum = 0.0; grade_wts = []; max_depth = 0
    for run in horse.runs:
        pr = raw_private_rating(run, horse.official_rating)
        if pr is None: continue
        capped = pr > ceiling; pr = min(pr, ceiling)
        rw = get_recency_weight(run.run_date, today)
        gw = get_grade_weight(run.grade_code)
        cm = run.signal.confidence_modifier if run.signal else 1.0
        cd = get_chain_discount(run.chain_depth)
        w  = rw * gw * cm
        run_ratings.append(RunRating(run=run, private_rating=round(pr, 2), capped=capped,
                                     recency_weight=rw, grade_weight=gw, confidence_mod=cm,
                                     chain_discount=cd, run_weight=round(w, 4)))
        total_weight += w; weighted_sum += pr * w; grade_wts.append(gw)
        max_depth = max(max_depth, run.chain_depth)
    weighted_avg = (weighted_sum / total_weight if total_weight > 0 else float(horse.official_rating))
    avg_gw = sum(grade_wts) / len(grade_wts) if grade_wts else 0.0
    return PrivateRatingResult(horse_name=horse.name, official_rating=horse.official_rating,
                               run_ratings=run_ratings, weighted_avg=round(weighted_avg, 2),
                               qualifying_runs=len(run_ratings), avg_grade_weight=round(avg_gw, 3),
                               max_chain_depth=max_depth, field_avg_or=field_avg_or,
                               field_ceiling=round(ceiling, 2))

def calculate_field_ratings(horses, today=None):
    avg_or, sd = field_stats(horses)
    results = [calculate_private_rating(h, avg_or, sd, today) for h in horses]
    return sorted(results, key=lambda r: r.weighted_avg, reverse=True)

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 4 — Preference Indices
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

GOING_MAP = {
    "firm":"FIRM","fast":"FIRM","good to firm":"GOOD_TO_FIRM","gf":"GOOD_TO_FIRM",
    "good":"GOOD","g":"GOOD","good to soft":"GOOD_TO_SOFT","gs":"GOOD_TO_SOFT",
    "yielding":"GOOD_TO_SOFT","soft":"SOFT","sft":"SOFT","s":"SOFT",
    "heavy":"HEAVY","hy":"HEAVY","h":"HEAVY",
}

GOING_ORDER = ["FIRM", "GOOD_TO_FIRM", "GOOD", "GOOD_TO_SOFT", "SOFT", "HEAVY"]

def normalise_going(raw_going_str):
    if not raw_going_str: return "UNKNOWN"
    s = raw_going_str.strip().lower()
    for key in sorted(GOING_MAP.keys(), key=len, reverse=True):
        if key in s: return GOING_MAP[key]
    return "UNKNOWN"

def going_adjacent(cat_a, cat_b):
    if cat_a == "UNKNOWN" or cat_b == "UNKNOWN": return False
    try:
        idx_a = GOING_ORDER.index(cat_a); idx_b = GOING_ORDER.index(cat_b)
        return abs(idx_a - idx_b) <= 1
    except ValueError:
        return False

COURSE_DATA = {
    "CHE": {"handed":"L","type":"GALLOPING","undulating":True,"hill":True},
    "ASC": {"handed":"R","type":"GALLOPING","undulating":False,"hill":False},
    "NEW": {"handed":"L","type":"GALLOPING","undulating":False,"hill":False},
    "SAN": {"handed":"R","type":"GALLOPING","undulating":False,"hill":True},
    "KEM": {"handed":"R","type":"SHARP","undulating":False,"hill":False},
    "HAY": {"handed":"L","type":"GALLOPING","undulating":False,"hill":False},
    "EXE": {"handed":"R","type":"GALLOPING","undulating":True,"hill":False},
    "TAU": {"handed":"R","type":"SHARP","undulating":False,"hill":False},
    "HUN": {"handed":"R","type":"SHARP","undulating":False,"hill":False},
    "WET": {"handed":"L","type":"GALLOPING","undulating":False,"hill":False},
    "DON": {"handed":"L","type":"GALLOPING","undulating":False,"hill":False},
    "YOR": {"handed":"L","type":"GALLOPING","undulating":False,"hill":False},
    "LEO": {"handed":"L","type":"GALLOPING","undulating":False,"hill":False},
    "PUN": {"handed":"R","type":"GALLOPING","undulating":False,"hill":False},
    "FAI": {"handed":"R","type":"GALLOPING","undulating":False,"hill":False},
    "NAV": {"handed":"L","type":"GALLOPING","undulating":False,"hill":False},
    "NAA": {"handed":"L","type":"SHARP","undulating":False,"hill":False},
    "GAL": {"handed":"R","type":"SHARP","undulating":True,"hill":False},
    "TIP": {"handed":"L","type":"SHARP","undulating":False,"hill":False},
    "LIM": {"handed":"R","type":"GALLOPING","undulating":False,"hill":False},
    "LIS": {"handed":"L","type":"SHARP","undulating":False,"hill":False},
    "UTT": {"handed":"L","type":"GALLOPING","undulating":True,"hill":False},
    "WOR": {"handed":"L","type":"SHARP","undulating":False,"hill":False},
    "CAR": {"handed":"R","type":"GALLOPING","undulating":True,"hill":True},
    "PER": {"handed":"R","type":"SHARP","undulating":False,"hill":False},
}

def course_match_score(course_a, course_b):
    a = COURSE_DATA.get(course_a.upper()); b = COURSE_DATA.get(course_b.upper())
    if not a or not b: return 0
    score = 0
    if a["handed"]     == b["handed"]:     score += 1
    if a["type"]       == b["type"]:       score += 1
    if a["undulating"] == b["undulating"]: score += 1
    if a["hill"]       == b["hill"]:       score += 1
    return score

def distance_band(furlongs):
    if furlongs < 16: return "SHORT"
    if furlongs < 20: return "MIDDLE"
    if furlongs < 24: return "EXTENDED"
    return "LONG"

PREF_MAX_SINGLE = 6.0
PREF_MAX_TOTAL  = 12.0

from dataclasses import dataclass

@dataclass
class PreferenceResult:
    horse_name:       str
    going_modifier:   float
    course_modifier:  float
    distance_modifier:float
    total_modifier:   float
    going_note:       str
    course_note:      str
    distance_note:    str

    def summary(self):
        lines = [
            f"  {self.horse_name} — Preference Modifiers",
            f"    Going    : {self.going_modifier:+.1f}lb  ({self.going_note})",
            f"    Course   : {self.course_modifier:+.1f}lb  ({self.course_note})",
            f"    Distance : {self.distance_modifier:+.1f}lb  ({self.distance_note})",
            f"    {'─'*29}",
            f"    TOTAL    : {self.total_modifier:+.1f}lb",
        ]
        return "\n".join(lines)

def calculate_preferences(horse, run_ratings, today_going, today_course, today_distance_f, min_confidence=1.00):
    today_band = distance_band(today_distance_f)
    going_runs = []; course_runs = []; distance_runs = []; all_runs = []
    for rr in run_ratings:
        run = rr.run
        if rr.run_weight <= 0: continue
        cm = run.signal.confidence_modifier if run.signal else 1.0
        if cm < min_confidence: continue
        pr = rr.private_rating; w = rr.run_weight; all_runs.append((pr, w))
        run_going = "UNKNOWN"
        if hasattr(run, 'going_category') and run.going_category: run_going = run.going_category
        if run_going != "UNKNOWN":
            if run_going == today_going: going_runs.append((pr, w))
            elif going_adjacent(run_going, today_going): going_runs.append((pr, w * 0.6))
        run_course = getattr(run, 'course_code', None)
        if run_course:
            match = course_match_score(run_course, today_course)
            if match >= 3: course_runs.append((pr, w * (match / 4.0)))
        run_distance_f = getattr(run, 'distance_f', None)
        if run_distance_f:
            run_band = distance_band(run_distance_f)
            if run_band == today_band: distance_runs.append((pr, w))

    def weighted_avg_from_pairs(pairs):
        if not pairs: return None
        total_w = sum(w for _, w in pairs)
        if total_w == 0: return None
        return sum(pr * w for pr, w in pairs) / total_w

    overall_avg = weighted_avg_from_pairs(all_runs)

    if len(going_runs) >= 1 and overall_avg is not None:
        going_avg = weighted_avg_from_pairs(going_runs)
        going_mod = going_avg - overall_avg; going_note = f"{today_going}, {len(going_runs)} run(s)"
    else:
        going_mod = 0.0; going_note = "insufficient data"

    if len(course_runs) >= 1 and overall_avg is not None:
        course_avg = weighted_avg_from_pairs(course_runs)
        course_mod = course_avg - overall_avg; course_note = f"match score >=3 for {today_course}, {len(course_runs)} run(s)"
    else:
        course_mod = 0.0; course_note = "insufficient data"

    if len(distance_runs) >= 1 and overall_avg is not None:
        dist_avg = weighted_avg_from_pairs(distance_runs)
        dist_mod = dist_avg - overall_avg; dist_note = f"{today_band}, {len(distance_runs)} run(s)"
    else:
        trip_signal_mod = 0.0; trip_notes = []
        for rr in run_ratings:
            if rr.run.signal:
                ts = rr.run.signal.trip_signal
                if ts.value == "WANTS_FURTHER": trip_signal_mod += 2.0 * rr.run_weight; trip_notes.append("WANTS_FURTHER signal")
                elif ts.value == "TRIP_TOO_FAR": trip_signal_mod -= 2.0 * rr.run_weight; trip_notes.append("TRIP_TOO_FAR signal")
        dist_mod = min(4.0, max(-4.0, trip_signal_mod))
        dist_note = f"{today_band}, no runs — " + (", ".join(set(trip_notes)) if trip_notes else "no trip signals")

    going_mod  = max(-PREF_MAX_SINGLE, min(PREF_MAX_SINGLE, going_mod))
    course_mod = max(-PREF_MAX_SINGLE, min(PREF_MAX_SINGLE, course_mod))
    dist_mod   = max(-PREF_MAX_SINGLE, min(PREF_MAX_SINGLE, dist_mod))

    total_raw = going_mod + course_mod + dist_mod
    if abs(total_raw) > PREF_MAX_TOTAL:
        scale = PREF_MAX_TOTAL / abs(total_raw)
        going_mod = round(going_mod * scale, 2); course_mod = round(course_mod * scale, 2)
        dist_mod = round(dist_mod * scale, 2); total_mod = going_mod + course_mod + dist_mod
    else:
        total_mod = total_raw

    return PreferenceResult(horse_name=horse.name, going_modifier=round(going_mod, 2),
                            course_modifier=round(course_mod, 2), distance_modifier=round(dist_mod, 2),
                            total_modifier=round(total_mod, 2), going_note=going_note,
                            course_note=course_note, distance_note=dist_note)

def calculate_field_preferences(horses, run_ratings_map, today_going, today_course, today_distance_f):
    results = {}
    for horse in horses:
        rr = run_ratings_map.get(horse.name, [])
        results[horse.name] = calculate_preferences(horse, rr, today_going, today_course, today_distance_f)
    return results

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 4 — Pace Classifier
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from dataclasses import dataclass, field
from collections import Counter

class PaceScenario:
    SLOWLY_RUN     = "SLOWLY_RUN"
    SOFT_LEAD      = "SOFT_LEAD"
    CONTESTED      = "CONTESTED"
    GENUINE_PACE   = "GENUINE_PACE"
    COLLAPSE_LIKELY= "COLLAPSE_LIKELY"

PACE_MODIFIERS = {
    "FRONT_RUNNER":  {
        PaceScenario.SLOWLY_RUN:1.10, PaceScenario.SOFT_LEAD:1.05,
        PaceScenario.CONTESTED:1.00, PaceScenario.GENUINE_PACE:1.00,
        PaceScenario.COLLAPSE_LIKELY:0.92},
    "PROMINENT": {
        PaceScenario.SLOWLY_RUN:1.05, PaceScenario.SOFT_LEAD:1.03,
        PaceScenario.CONTESTED:1.03, PaceScenario.GENUINE_PACE:1.02,
        PaceScenario.COLLAPSE_LIKELY:0.96},
    "MIDFIELD": {
        PaceScenario.SLOWLY_RUN:1.00, PaceScenario.SOFT_LEAD:1.00,
        PaceScenario.CONTESTED:1.00, PaceScenario.GENUINE_PACE:1.03,
        PaceScenario.COLLAPSE_LIKELY:1.02},
    "HOLD_UP": {
        PaceScenario.SLOWLY_RUN:0.92, PaceScenario.SOFT_LEAD:0.95,
        PaceScenario.CONTESTED:0.95, PaceScenario.GENUINE_PACE:1.08,
        PaceScenario.COLLAPSE_LIKELY:1.06},
    "UNKNOWN": {
        PaceScenario.SLOWLY_RUN:1.00, PaceScenario.SOFT_LEAD:1.00,
        PaceScenario.CONTESTED:1.00, PaceScenario.GENUINE_PACE:1.00,
        PaceScenario.COLLAPSE_LIKELY:1.00},
}

@dataclass
class HorsePaceProfile:
    horse_name:       str
    assigned_style:   str
    style_confidence: float
    slow_history:     bool
    pace_modifier:    float
    scenario:         str
    scenario_prob:    float
    note:             str

    def summary(self):
        return (f"  {self.horse_name:<22} style={self.assigned_style:<14} "
                f"modifier=x{self.pace_modifier:.3f}  ({self.scenario} @ {self.scenario_prob:.0%})")

@dataclass
class PaceResult:
    scenario:          str
    scenario_prob:     float
    front_runner_count:int
    field_size:        int
    profiles:          list
    pace_map_note:     str

    def summary(self):
        lines = [
            f"\n  PACE SCENARIO : {self.scenario}  ({self.scenario_prob:.0%} probability)",
            f"  Front runners : {self.front_runner_count} of {self.field_size}",
            f"  Note          : {self.pace_map_note}",
            f"  {'─'*65}",
        ]
        for p in self.profiles: lines.append(p.summary())
        return "\n".join(lines)

def assign_running_style(horse, run_ratings, max_comments=3):
    signals = [rr.run.signal for rr in run_ratings if rr.run.signal is not None][:max_comments]
    if not signals: return "UNKNOWN", 0.5
    modal = modal_running_style(signals, max_comments); style_str = modal.value
    styles = [s.running_style.value for s in signals if s.running_style.value != "UNKNOWN"]
    if not styles: return "UNKNOWN", 0.5
    counts = Counter(styles); top_count = counts.most_common(1)[0][1]; total = len(styles)
    if top_count == total: confidence = 1.00
    elif top_count / total >= 0.67: confidence = 0.85
    else: confidence = 0.70
    return style_str, confidence

def detect_slow_history(run_ratings):
    front_runs = 0; headed_early = 0
    for rr in run_ratings:
        sig = rr.run.signal
        if sig is None: continue
        if sig.running_style.value != "FRONT_RUNNER": continue
        front_runs += 1
        if sig.trip_signal.value == "TRIP_TOO_FAR": headed_early += 1
    if front_runs == 0: return False
    return (headed_early / front_runs) >= 0.5

def determine_scenario(front_runner_count, field_size, has_slow_history):
    fr = front_runner_count
    if fr == 0: return PaceScenario.SLOWLY_RUN, 0.85
    if fr == 1:
        if has_slow_history: return PaceScenario.SOFT_LEAD, 0.70
        else: return PaceScenario.CONTESTED, 0.50
    if fr in (2, 3):
        prob = 0.75 if field_size >= 16 else 0.65
        return PaceScenario.GENUINE_PACE, prob
    return PaceScenario.COLLAPSE_LIKELY, 0.80

def apply_pace_modifier(rating, style_str, scenario, scenario_prob):
    modifier = PACE_MODIFIERS.get(style_str, PACE_MODIFIERS["UNKNOWN"])
    mod_val  = modifier.get(scenario, 1.00)
    blended  = 1.0 + (mod_val - 1.0) * scenario_prob
    return round(rating * blended, 2), round(blended, 4)

def classify_pace(horses, run_ratings_map, field_size=None):
    if field_size is None: field_size = len(horses)
    profiles_tmp = []
    for horse in horses:
        rr = run_ratings_map.get(horse.name, [])
        style, conf = assign_running_style(horse, rr)
        slow = detect_slow_history(rr) if style == "FRONT_RUNNER" else False
        profiles_tmp.append({"horse":horse,"style":style,"confidence":conf,"slow":slow})
    front_runners = [p for p in profiles_tmp if p["style"] == "FRONT_RUNNER"]
    fr_count = len(front_runners)
    sole_slow = (fr_count == 1 and front_runners[0]["slow"] if fr_count == 1 else False)
    scenario, prob = determine_scenario(fr_count, field_size, sole_slow)
    fr_names = [p["horse"].name for p in front_runners]
    note = f"Front runner(s): {', '.join(fr_names)}" if fr_names else "No identified front runners — slowly run likely"
    profiles = []
    for p in profiles_tmp:
        mod_val = PACE_MODIFIERS.get(p["style"], PACE_MODIFIERS["UNKNOWN"])
        raw_mod = mod_val.get(scenario, 1.00)
        blended = round(1.0 + (raw_mod - 1.0) * prob, 4)
        profiles.append(HorsePaceProfile(horse_name=p["horse"].name, assigned_style=p["style"],
                                         style_confidence=p["confidence"], slow_history=p["slow"],
                                         pace_modifier=blended, scenario=scenario, scenario_prob=prob,
                                         note=f"style confidence {p['confidence']:.0%}"))
    return PaceResult(scenario=scenario, scenario_prob=prob, front_runner_count=fr_count,
                      field_size=field_size, profiles=profiles, pace_map_note=note)

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 5 — Dampening Engine
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from dataclasses import dataclass, field as dc_field
import math

D1_CONFIDENCE = {6:1.00, 5:0.95, 4:0.88, 3:0.78, 2:0.65, 1:0.50, 0:0.00}

def d1_confidence(qualifying_runs):
    return D1_CONFIDENCE.get(min(qualifying_runs, 6), 1.00)

def d7_regression(meaningful_runs):
    if meaningful_runs >= 3: return 0.0
    return round((3 - meaningful_runs) / 3, 4)

def band_contributions(qualifying_runs, avg_grade_weight, max_chain_depth,
                       stepping_up_trip, stepping_up_class, max_pref_modifier):
    contribs = {}
    missing = max(0, 6 - qualifying_runs)
    if missing > 0: contribs["low_runs"] = min(18, missing * 6)
    chain_extra = max(0, max_chain_depth - 0)
    if chain_extra > 0: contribs["chain_depth"] = min(12, chain_extra * 4)
    if avg_grade_weight < 0.60: contribs["low_grade"] = 4
    if stepping_up_trip: contribs["trip_step_up"] = 4
    if stepping_up_class: contribs["class_step_up"] = 4
    if max_pref_modifier > 4.0: contribs["high_pref_mod"] = 2
    if contribs: band = math.sqrt(sum(v ** 2 for v in contribs.values()))
    else: band = 2.0
    return contribs, round(band, 1)

def band_label(band_width):
    if band_width <= 3: return "NARROW"
    if band_width <= 7: return "MEDIUM"
    return "WIDE"

@dataclass
class DampeningAudit:
    d1_input:float; d1_output:float; d1_confidence:float; d1_field_avg:float
    d3_output:float; d3_note:str; d4_output:float; d4_note:str
    d5_output:float; d5_going:float; d5_course:float; d5_distance:float
    d7_output:float; d7_regression:float
    pace_output:float; pace_modifier:float; pace_note:str

@dataclass
class DampeningResult:
    horse_name:str; official_rating:int; private_rating:float; projected_rating:float
    band_low:float; band_high:float; band_width:float; band_label:str
    band_contributions:dict; audit:DampeningAudit; field_avg_or:float

    def summary(self):
        lines = [
            f"\n{'═'*60}",
            f"  {self.horse_name}  (OR {self.official_rating})",
            f"{'─'*60}",
            f"  Private rating (L3) : {self.private_rating:.1f}",
            f"  Projected rating    : {self.projected_rating:.1f}",
            f"  Confidence band     : {self.band_low:.1f} — {self.band_high:.1f}  (+-{self.band_width:.1f}lb, {self.band_label})",
            f"{'─'*60}",
            f"  DAMPENING AUDIT:",
            f"    D1 volume     : {self.audit.d1_input:.1f} -> {self.audit.d1_output:.1f}  (conf {self.audit.d1_confidence:.2f}, field avg {self.audit.d1_field_avg:.1f})",
            f"    D3 chain      : -> {self.audit.d3_output:.1f}  ({self.audit.d3_note})",
            f"    D4 outlier    : -> {self.audit.d4_output:.1f}  ({self.audit.d4_note})",
            f"    D5 prefs      : -> {self.audit.d5_output:.1f}  (going {self.audit.d5_going:+.1f} / course {self.audit.d5_course:+.1f} / dist {self.audit.d5_distance:+.1f})",
            f"    D7 regression : -> {self.audit.d7_output:.1f}  (strength {self.audit.d7_regression:.2f})",
            f"    Pace modifier : -> {self.audit.pace_output:.1f}  (x{self.audit.pace_modifier:.4f}, {self.audit.pace_note})",
            f"{'─'*60}",
            f"  Band contributors : {self.band_contributions}",
            f"{'═'*60}",
        ]
        return "\n".join(lines)

def dampen(horse, rating_result, pref_result, pace_profile, field_avg_or,
           field_avg_projected=None, stepping_up_trip=False, stepping_up_class=False):
    pr = rating_result.weighted_avg; qr = rating_result.qualifying_runs
    gw = rating_result.avg_grade_weight; cd = rating_result.max_chain_depth
    ceil = rating_result.field_ceiling
    conf_d1 = d1_confidence(qr)
    d1_out = (pr * conf_d1) + (field_avg_or * (1.0 - conf_d1))
    d3_in = d1_out
    if cd == 0: d3_out = d3_in; d3_note = "all direct runs, no chain adjustment"
    else: d3_out = d3_in; d3_note = f"chain depth {cd} applied in Cell 2 run weights"
    capped_runs = [rr for rr in rating_result.run_ratings if rr.capped]
    if capped_runs: d4_note = f"{len(capped_runs)} run(s) capped at ceiling {ceil:.1f}"
    else: d4_note = f"no runs exceeded ceiling {ceil:.1f}"
    d4_out = d3_out
    going_mod = pref_result.going_modifier; course_mod = pref_result.course_modifier
    dist_mod = pref_result.distance_modifier
    d5_out = d4_out + going_mod + course_mod + dist_mod
    d6_out = d5_out
    meaningful = sum(1 for rr in rating_result.run_ratings if rr.grade_weight >= 0.55)
    reg_strength = d7_regression(meaningful)
    d7_out = (d6_out * (1.0 - reg_strength)) + (field_avg_or * reg_strength)
    pace_mod = pace_profile.pace_modifier if pace_profile else 1.0
    pace_out = round(d7_out * pace_mod, 2)
    pace_note = (f"{pace_profile.assigned_style} / {pace_profile.scenario} @ {pace_profile.scenario_prob:.0%}"
                 if pace_profile else "no pace profile")
    max_pref = max(abs(going_mod), abs(course_mod), abs(dist_mod))
    contribs, band = band_contributions(qualifying_runs=qr, avg_grade_weight=gw, max_chain_depth=cd,
                                        stepping_up_trip=stepping_up_trip, stepping_up_class=stepping_up_class,
                                        max_pref_modifier=max_pref)
    projected = pace_out; band_lo = round(projected - band, 1); band_hi = round(projected + band, 1)
    audit = DampeningAudit(d1_input=round(pr,2), d1_output=round(d1_out,2), d1_confidence=conf_d1,
                           d1_field_avg=round(field_avg_or,2), d3_output=round(d3_out,2), d3_note=d3_note,
                           d4_output=round(d4_out,2), d4_note=d4_note, d5_output=round(d5_out,2),
                           d5_going=going_mod, d5_course=course_mod, d5_distance=dist_mod,
                           d7_output=round(d7_out,2), d7_regression=reg_strength,
                           pace_output=pace_out, pace_modifier=pace_mod, pace_note=pace_note)
    return DampeningResult(horse_name=horse.name, official_rating=horse.official_rating,
                           private_rating=round(pr,2), projected_rating=projected,
                           band_low=band_lo, band_high=band_hi, band_width=band,
                           band_label=band_label(band), band_contributions=contribs,
                           audit=audit, field_avg_or=round(field_avg_or,2))

def dampen_field(horses, rating_results_map, pref_results_map, pace_result, field_avg_or,
                 step_up_trip_map=None, step_up_class_map=None):
    pace_map = {p.horse_name: p for p in pace_result.profiles}
    step_trip_map = step_up_trip_map or {}; step_class_map = step_up_class_map or {}
    results = {}
    for horse in horses:
        rr = rating_results_map.get(horse.name); pr = pref_results_map.get(horse.name)
        pace = pace_map.get(horse.name)
        if rr is None: continue
        if pr is None:
            pr = PreferenceResult(horse_name=horse.name, going_modifier=0.0, course_modifier=0.0,
                                  distance_modifier=0.0, total_modifier=0.0,
                                  going_note="n/a", course_note="n/a", distance_note="n/a")
        results[horse.name] = dampen(horse=horse, rating_result=rr, pref_result=pr, pace_profile=pace,
                                     field_avg_or=field_avg_or,
                                     stepping_up_trip=step_trip_map.get(horse.name, False),
                                     stepping_up_class=step_class_map.get(horse.name, False))
    return dict(sorted(results.items(), key=lambda x: x[1].projected_rating, reverse=True))

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 6 — Market-Anchored Value Calculator
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from dataclasses import dataclass
import math

BASE_EDGE_SCALE = 20.0
BASE_DISTANCE_F = 16.0

def edge_scale_for_distance(distance_f):
    if not distance_f or distance_f <= 0: return BASE_EDGE_SCALE
    return BASE_EDGE_SCALE * (BASE_DISTANCE_F / distance_f)

MAX_ADJ_MULTIPLIER = 6.0
CONFIDENCE_SCALE = {'NARROW':1.00, 'MEDIUM':0.70, 'WIDE':0.40}
MAX_GAP_LEADER = {'G1':20.0, 'G2':22.0, 'HCAP':28.0, 'DEFAULT':25.0}
MIN_VALUE_RATIO = 1.5
MIN_KELLY_PCT_SINGLE = 0.5
MIN_KELLY_PCT_BASKET = 0.3
BASKET_MIN_ODDS = 10.0
BASKET_MAX_SIZE = 4
BASKET_MIN_COLLECTIVE = 1.0
KELLY_FRACTION = 0.25
MAX_KELLY_PCT = 5.0

def place_terms(field_size):
    if field_size <= 4:  return None, 0
    if field_size <= 7:  return 2.5, 2
    if field_size <= 11: return 3.0, 3
    if field_size <= 15: return 3.5, 3
    return 4.0, 4

def fractional_to_decimal(fractional_str):
    if not fractional_str: return None
    s = str(fractional_str).strip().lower()
    if s in ("evens","evs","1/1"): return 2.0
    if "/" in s:
        parts = s.split("/")
        try:
            num = float(parts[0].strip()); den = float(parts[1].strip())
            return round((num / den) + 1.0, 4)
        except (ValueError, ZeroDivisionError): return None
    try: return float(s)
    except ValueError: return None

def decimal_to_fractional(decimal_odds):
    common = {
        1.25:"1/4",1.33:"1/3",1.5:"1/2",1.67:"4/6",1.8:"4/5",2.0:"evens",
        2.5:"6/4",3.0:"2/1",3.5:"5/2",4.0:"3/1",4.5:"7/2",5.0:"4/1",
        5.5:"9/2",6.0:"5/1",7.0:"6/1",8.0:"7/1",9.0:"8/1",10.0:"9/1",
        11.0:"10/1",13.0:"12/1",14.0:"13/1",15.0:"14/1",16.0:"15/1",
        17.0:"16/1",21.0:"20/1",26.0:"25/1",29.0:"28/1",31.0:"30/1",
        34.0:"33/1",41.0:"40/1",51.0:"50/1",67.0:"66/1",101.0:"100/1",
    }
    rounded = round(decimal_odds, 2)
    if rounded in common: return common[rounded]
    closest = min(common.keys(), key=lambda k: abs(k - decimal_odds))
    if abs(closest - decimal_odds) < 1.5: return common[closest]
    return f"{decimal_odds - 1:.0f}/1"

def adjust_probability(market_prob, true_edge, band_label, distance_f=16.0):
    conf = CONFIDENCE_SCALE.get(band_label, 0.40)
    e_scale = edge_scale_for_distance(distance_f)
    raw_mult = 2.0 ** (true_edge / e_scale)
    eff_mult = 1.0 + (raw_mult - 1.0) * conf
    eff_mult = max(1.0 / MAX_ADJ_MULTIPLIER, min(MAX_ADJ_MULTIPLIER, eff_mult))
    return market_prob * eff_mult

def kelly_stake_pct(model_prob, decimal_odds):
    b = decimal_odds
    if b <= 1.0: return 0.0
    kelly_full = (model_prob * (b - 1.0) - (1.0 - model_prob)) / (b - 1.0)
    kelly_frac = kelly_full * KELLY_FRACTION
    return round(min(max(kelly_frac, 0.0), MAX_KELLY_PCT / 100.0) * 100.0, 2)

def kelly_to_label(kelly_pct):
    if kelly_pct <= 0:   return "NO BET"
    if kelly_pct >= 3.0: return "FULL"
    if kelly_pct >= 1.5: return "HALF"
    if kelly_pct >= 0.5: return "QUARTER"
    return "SMALL"

@dataclass
class EVResult:
    horse_name:str; official_rating:int; projected_rating:float
    band_low:float; band_high:float; band_label:str
    field_avg_projected:float; field_avg_or:float
    individual_edge:float; official_edge:float; true_edge:float; edge_confidence:str
    market_prob:float; model_prob:float; model_price_dec:float; model_price_frac:str
    value_ratio:float; decimal_odds:float; fractional_odds:str
    kelly_pct:float; kelly_ew_pct:float; stake:str
    place_factor:float; places_paid:int; action:str
    manual_override:bool; override_note:str; or_proxied:bool; proxy_note:str
    gap_to_leader:float; field_rank:int

    def summary(self):
        return "\n".join([
            f"  {self.horse_name}  (OR {self.official_rating})",
            f"    Projected   : {self.projected_rating:.1f}  [{self.band_low:.1f}--{self.band_high:.1f}, {self.band_label}]",
            f"    True Edge   : {self.true_edge:+.1f}lb  ({self.edge_confidence} confidence)",
            f"    Market      : {self.fractional_odds}  (implied {self.market_prob:.1%})",
            f"    Model price : {self.model_price_frac}  (adjusted {self.model_prob:.1%})",
            f"    Value ratio : {self.value_ratio:.2f}x  ({'VALUE' if self.value_ratio >= MIN_VALUE_RATIO else 'no value'})",
            f"    Kelly stake : {self.kelly_pct:.2f}% of bank  [{self.stake}]",
            f"    Action      : {self.action}",
        ])

@dataclass
class RaceEVSummary:
    race_name:str; field_size:int; field_avg_or:float; field_avg_projected:float
    field_true_edge_avg:float; results:list; pace_scenario:str; pace_prob:float
    selection_mode:str; basket_collective_ev:float

    def summary(self):
        lines = [
            f"\n{'═'*72}", f"  {self.race_name.upper()}",
            f"  Field avg OR: {self.field_avg_or:.1f}  | Avg projected: {self.field_avg_projected:.1f}  | Pace: {self.pace_scenario} ({self.pace_prob:.0%})",
            f"{'─'*72}",
            f"  {'Horse':<22} {'OR':>4} {'Proj':>6} {'Mkt':>8} {'Model':>7} {'Val x':>5} {'Kelly%':>6} {'Action':<10}",
            f"{'─'*72}",
        ]
        for r in self.results:
            if r.manual_override:
                lines.append(f"  {r.horse_name:<22} {r.official_rating:>4}  WATCH -- {r.override_note}")
                continue
            lines.append(f"  {r.horse_name:<22} {r.official_rating:>4} {r.projected_rating:>6.1f} "
                         f"{r.fractional_odds:>8} {r.model_price_frac:>7} {r.value_ratio:>5.2f}x "
                         f"{r.kelly_pct:>6.2f}% {r.action:<10}")
        lines.append(f"{'═'*72}")
        return "\n".join(lines)

def calculate_ev(dampening_results, odds_map, race_name, field_avg_or, pace_result,
                 override_map=None, proxied_or_map=None, race_grade='DEFAULT', distance_f=16.0):
    override_map = override_map or {}; proxied_or_map = proxied_or_map or {}
    field_size = len(dampening_results); pf, places = place_terms(field_size)
    projected_ratings = [dr.projected_rating for dr in dampening_results.values()]
    field_avg_proj = sum(projected_ratings) / len(projected_ratings)
    leader_projected = max(projected_ratings)
    sorted_proj = sorted(dampening_results.items(), key=lambda kv: kv[1].projected_rating, reverse=True)
    field_rank_map = {name: i+1 for i, (name, _) in enumerate(sorted_proj)}
    max_gap = MAX_GAP_LEADER.get(race_grade, MAX_GAP_LEADER['DEFAULT'])
    true_edges = {}
    for name, dr in dampening_results.items():
        ind_edge = dr.projected_rating - field_avg_proj
        off_edge = dr.official_rating - field_avg_or
        true_edges[name] = ind_edge - off_edge
    field_te_avg = sum(true_edges.values()) / len(true_edges)
    dec_odds_map = {}; raw_market_probs = {}; adj_probs = {}
    for name, dr in dampening_results.items():
        odds_str = odds_map.get(name, "")
        dec_odds = fractional_to_decimal(odds_str) if odds_str else None
        dec_odds_map[name] = dec_odds
        mkt_prob = (1.0 / dec_odds) if (dec_odds and dec_odds > 1.0) else (1.0 / field_size)
        raw_market_probs[name] = mkt_prob
        adj_probs[name] = adjust_probability(mkt_prob, true_edges[name], dr.band_label, distance_f)
    total_adj = sum(adj_probs.values())
    model_probs = {name: p / total_adj for name, p in adj_probs.items()}
    results = []
    for name, dr in dampening_results.items():
        te = true_edges[name]; dec_odds = dec_odds_map[name]
        odds_str = odds_map.get(name, "n/a"); model_prob = model_probs[name]
        mkt_prob = raw_market_probs[name]
        model_price_dec = 1.0 / model_prob if model_prob > 0 else 999.0
        model_price_frac = decimal_to_fractional(model_price_dec)
        value_ratio = (dec_odds / model_price_dec) if (dec_odds and dec_odds > 1.0) else 0.0
        edge_conf = {"NARROW":"HIGH","MEDIUM":"MEDIUM","WIDE":"LOW"}.get(dr.band_label,"LOW")
        kelly_win = kelly_stake_pct(model_prob, dec_odds) if dec_odds else 0.0
        kelly_ew = 0.0
        if dec_odds and pf and dec_odds >= 6.0:
            place_prob = model_prob * pf; place_dec = ((dec_odds - 1.0) / 4.0) + 1.0
            kelly_ew = kelly_stake_pct(place_prob, place_dec)
        gap_leader = dr.projected_rating - leader_projected
        rank = field_rank_map[name]; in_contention = (gap_leader >= -max_gap)
        if name in override_map or not dec_odds:
            action = "WATCH"; stake = "NO BET"; kelly_win = 0.0
        elif value_ratio >= MIN_VALUE_RATIO and kelly_win > 0 and in_contention:
            if kelly_ew > 0 and dec_odds >= 10.0: action = "CANDIDATE_EW"; stake = kelly_to_label(kelly_ew)
            else: action = "CANDIDATE_WIN"; stake = kelly_to_label(kelly_win)
        elif value_ratio >= 1.2 and kelly_win > 0 and in_contention:
            action = "CONSIDER"; stake = "NO BET"
        elif te > 5 and dr.band_label == "NARROW" and in_contention:
            action = "CONSIDER"; stake = "NO BET"
        elif value_ratio < 0.7 or te < -10:
            action = "AVOID"; stake = "NO BET"
        else:
            action = "WATCH"; stake = "NO BET"
        results.append(EVResult(
            horse_name=name, official_rating=dr.official_rating, projected_rating=dr.projected_rating,
            band_low=dr.band_low, band_high=dr.band_high, band_label=dr.band_label,
            field_avg_projected=round(field_avg_proj,2), field_avg_or=round(field_avg_or,2),
            individual_edge=round(dr.projected_rating - field_avg_proj, 2),
            official_edge=round(dr.official_rating - field_avg_or, 2),
            true_edge=round(te,2), edge_confidence=edge_conf,
            market_prob=round(mkt_prob,4), model_prob=round(model_prob,4),
            model_price_dec=round(model_price_dec,2), model_price_frac=model_price_frac,
            value_ratio=round(value_ratio,3), decimal_odds=dec_odds or 0.0, fractional_odds=odds_str,
            kelly_pct=kelly_win, kelly_ew_pct=kelly_ew, stake=stake,
            place_factor=pf or 0.0, places_paid=places, action=action,
            manual_override=name in override_map, override_note=override_map.get(name,""),
            or_proxied=name in proxied_or_map, proxy_note=proxied_or_map.get(name,""),
            gap_to_leader=round(gap_leader,1), field_rank=rank))
    results.sort(key=lambda r: r.projected_rating, reverse=True)
    candidates = [r for r in results if r.action.startswith("CANDIDATE")]
    candidates_by_value = sorted(candidates, key=lambda r: r.value_ratio, reverse=True)
    selection_mode = "NONE"; selected_names = set()
    if candidates_by_value:
        best = candidates_by_value[0]
        best_kelly = best.kelly_ew_pct if best.kelly_ew_pct and best.kelly_ew_pct > 0 else best.kelly_pct
        if best.value_ratio >= MIN_VALUE_RATIO and best_kelly >= MIN_KELLY_PCT_SINGLE:
            selection_mode = "SINGLE"; selected_names = {best.horse_name}
        longshot_candidates = [r for r in candidates_by_value
                               if r.decimal_odds >= BASKET_MIN_ODDS and r.value_ratio >= MIN_VALUE_RATIO
                               and (r.kelly_ew_pct or r.kelly_pct or 0) >= MIN_KELLY_PCT_BASKET]
        if len(longshot_candidates) >= 2 and best.decimal_odds >= BASKET_MIN_ODDS:
            basket = longshot_candidates[:BASKET_MAX_SIZE]
            collective_ev = sum(r.model_prob * r.decimal_odds for r in basket) / len(basket)
            if collective_ev > BASKET_MIN_COLLECTIVE:
                selection_mode = "BASKET"; selected_names = {r.horse_name for r in basket}
    for r in results:
        if r.action.startswith("CANDIDATE"):
            base_action = "EW" if "EW" in r.action else "WIN"
            if r.horse_name in selected_names: r.action = base_action
            else: r.action = "CONSIDER"; r.stake = "NO BET"
    return RaceEVSummary(
        race_name=race_name, field_size=field_size,
        field_avg_or=round(field_avg_or,2), field_avg_projected=round(field_avg_proj,2),
        field_true_edge_avg=round(field_te_avg,2), results=results,
        pace_scenario=pace_result.scenario, pace_prob=pace_result.scenario_prob,
        selection_mode=selection_mode,
        basket_collective_ev=round(
            sum(r.model_prob * r.decimal_odds for r in results if r.action in ("WIN","EW")) /
            max(len([r for r in results if r.action in ("WIN","EW")]), 1), 3))

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 7 — Race Output Formatter
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from datetime import datetime

def _bar(char="═", width=65): return char * width
def _divider(width=65): return "─" * width

def _action_label(action, stake):
    labels = {"WIN":"▶ WIN","EW":"▶ EACH WAY","CONSIDER":"◑ CONSIDER","AVOID":"✕ AVOID","WATCH":"◎ WATCH"}
    label = labels.get(action, action)
    if stake and stake != "NO BET" and action not in ("AVOID","WATCH","CONSIDER"):
        return f"{label}  [{stake} STAKE]"
    return label

def _confidence_label(band_label, true_edge):
    if band_label == "NARROW" and abs(true_edge) >= 5: return "Strong signal — well evidenced"
    if band_label == "NARROW": return "Reliable evidence — modest edge"
    if band_label == "MEDIUM" and abs(true_edge) >= 5: return "Reasonable signal — some uncertainty"
    if band_label == "MEDIUM": return "Limited evidence — treat cautiously"
    if abs(true_edge) >= 8: return "Large edge claim — thin evidence, small stake only"
    return "Insufficient evidence — speculative only"

def full_report(ev_summary, dampening_map=None, race_date=None):
    lines = []
    date_str = race_date or datetime.today().strftime("%d %B %Y")
    lines.append(_bar())
    lines.append(f"  PRIVATE RATINGS ENGINE — RACE REPORT")
    lines.append(f"  {ev_summary.race_name.upper()}")
    lines.append(f"  {date_str}  |  Field: {ev_summary.field_size} runners")
    lines.append(f"  Field avg OR: {ev_summary.field_avg_or:.1f}  |  Avg projected: {ev_summary.field_avg_projected:.1f}  |  True edge avg: {ev_summary.field_true_edge_avg:+.1f}lb")
    lines.append(f"  Pace: {ev_summary.pace_scenario}  ({ev_summary.pace_prob:.0%} probability)")
    lines.append(_bar()); lines.append("")
    for r in ev_summary.results:
        name_display = r.horse_name + (" *" if r.or_proxied else "")
        lines.append(_divider())
        lines.append(f"  #{name_display}  |  OR {r.official_rating}{'*' if r.or_proxied else ''}  |  {_action_label(r.action, r.stake)}")
        lines.append(_divider())
        if r.manual_override: lines.append(f"  ◎ WATCH — {r.override_note}"); lines.append(""); continue
        if r.or_proxied: lines.append(f"  OR proxied from RPR — {r.proxy_note}"); lines.append("")
        lines.append(f"  {'Projected rating':<22}: {r.projected_rating:.1f}")
        lines.append(f"  {'Confidence band':<22}: {r.band_low:.1f} — {r.band_high:.1f}  (+-{r.band_high - r.projected_rating:.1f}lb, {r.band_label})")
        lines.append(f"  {'Evidence quality':<22}: {_confidence_label(r.band_label, r.true_edge)}")
        lines.append("")
        lines.append(f"  {'Individual edge':<22}: {r.individual_edge:+.1f}lb  (projected vs field avg {r.field_avg_projected:.1f})")
        lines.append(f"  {'Official edge':<22}: {r.official_edge:+.1f}lb  (OR vs field avg OR {r.field_avg_or:.1f})")
        lines.append(f"  {'TRUE EDGE':<22}: {r.true_edge:+.1f}lb  [{r.edge_confidence} confidence]")
        lines.append("")
        if hasattr(r, 'model_prob'):
            lines.append(f"  {'Market price':<22}: {r.fractional_odds}  (implied {r.market_prob:.1%})")
            lines.append(f"  {'Model price':<22}: {r.model_price_frac}  (adjusted {r.model_prob:.1%})")
            v_label = "VALUE" if r.value_ratio >= 1.5 else ("marginal" if r.value_ratio >= 1.2 else "no value")
            lines.append(f"  {'Value ratio':<22}: {r.value_ratio:.2f}x  [{v_label}]")
            lines.append(f"  {'Kelly stake':<22}: {r.kelly_pct:.2f}% of bank  [{r.stake}]")
            if r.kelly_ew_pct and r.kelly_ew_pct > 0:
                lines.append(f"  {'Kelly EW stake':<22}: {r.kelly_ew_pct:.2f}% of bank  ({r.places_paid} places)")
        else:
            lines.append(f"  {'Market odds':<22}: {r.fractional_odds}  (decimal {r.decimal_odds:.2f})")
        lines.append("")
        if dampening_map and r.horse_name in dampening_map:
            dr = dampening_map[r.horse_name]; a = dr.audit
            lines.append(f"  DAMPENING TRAIL:")
            lines.append(f"    Private rating (L3)  : {dr.private_rating:.1f}")
            lines.append(f"    D1 volume ({dr.audit.d1_confidence:.0%} conf)  : {a.d1_input:.1f} -> {a.d1_output:.1f}")
            lines.append(f"    D5 preferences       : going {a.d5_going:+.1f} / course {a.d5_course:+.1f} / dist {a.d5_distance:+.1f}")
            lines.append(f"    D7 regression        : strength {a.d7_regression:.2f}  -> {a.d7_output:.1f}")
            lines.append(f"    Pace modifier        : x{a.pace_modifier:.4f}  ({a.pace_note})")
            if dr.band_contributions: lines.append(f"    Band contributors    : {dr.band_contributions}")
            lines.append("")
        lines.append("")
    lines.append(_bar()); lines.append("  SUMMARY TABLE  (sorted by True Edge)"); lines.append(_bar())
    lines.append(f"  {'Horse':<22} {'OR':>4} {'Proj':>6} {'Band':>11} {'Edge':>7} {'WinP':>6} {'Odds':>8} {'EV%':>6}  {'Action'}")
    lines.append(_divider())
    for r in ev_summary.results:
        name_display = r.horse_name + ("*" if r.or_proxied else "")
        if r.manual_override: lines.append(f"  {name_display:<22} {r.official_rating:>4}  WATCH"); continue
        band_str = f"{r.band_low:.0f}--{r.band_high:.0f}"
        lines.append(f"  {name_display:<22} {r.official_rating:>4} {r.projected_rating:>6.1f} {band_str:>11} {r.true_edge:>+7.1f} {r.model_prob:>6.1%} {r.fractional_odds:>8} {r.value_ratio:>5.2f}x  {r.kelly_pct:>5.2f}%  {_action_label(r.action, r.stake)}")
    lines.append(_bar())
    primary = [r for r in ev_summary.results if r.action in ("WIN","EW") and r.stake != "NO BET"]
    consider = [r for r in ev_summary.results if r.action == "CONSIDER"]
    avoid = [r for r in ev_summary.results if r.action == "AVOID"]
    if primary or consider or avoid:
        lines.append(""); lines.append("  SELECTIONS"); lines.append(_divider())
        if primary:
            for r in primary:
                nd = r.horse_name + ("*" if r.or_proxied else "")
                lines.append(f"  ▶ {nd:<20} {r.fractional_odds:<8}  {r.action}  [{r.stake} STAKE]  val {r.value_ratio:.2f}x  Kelly {r.kelly_pct:.2f}%")
        if consider:
            lines.append(""); lines.append("  ◑ CONSIDER — strong form signal, price too short for positive EV:")
            for r in consider:
                nd = r.horse_name + ("*" if r.or_proxied else "")
                lines.append(f"    {nd:<20} {r.fractional_odds:<8}  Edge {r.true_edge:+.1f}lb  Band {r.band_label}  val {r.value_ratio:.2f}x  Kelly {r.kelly_pct:.2f}%")
        if avoid:
            lines.append("")
            for r in avoid:
                nd = r.horse_name + ("*" if r.or_proxied else "")
                lines.append(f"  ✕ AVOID: {nd:<18} {r.fractional_odds:<8}  edge {r.true_edge:+.1f}lb  val {r.value_ratio:.2f}x")
        lines.append(_divider())
    lines.append(""); lines.append(f"  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"); lines.append(_bar())
    return "\n".join(lines)

def quick_card(ev_summary):
    ranked = sorted(ev_summary.results, key=lambda r: r.projected_rating, reverse=True)
    lines = [
        f"QUICK CARD — {ev_summary.race_name.upper()}",
        f"Pace: {ev_summary.pace_scenario} ({ev_summary.pace_prob:.0%})  | Field avg OR {ev_summary.field_avg_or:.0f}  | {len(ranked)} runners",
        _divider(78),
        f"  {'#':>3}  {'Horse':<23} {'OR':>4}  {'Proj':>5}  {'Gap':>5}  {'Mkt':>7}  {'Model':>7}  {'Val':>5}  {'K%':>4}  Action",
        _divider(78),
    ]
    proxied_names = []
    for i, r in enumerate(ranked):
        name_str = r.horse_name + ("*" if r.or_proxied else "")
        gap_str = f"{r.gap_to_leader:+.0f}" if hasattr(r, 'gap_to_leader') else "  --"
        ev_val = r.kelly_ew_pct if r.kelly_ew_pct is not None else r.kelly_pct
        ev_str = f"{ev_val:+.0f}%" if ev_val else "  --"
        if r.manual_override: flag = " ◎"; action_str = "WATCH"
        else:
            flag = ("▶▶" if r.action in ("WIN","EW") and r.stake=="FULL" else
                    " ▶" if r.action in ("WIN","EW") and r.stake=="HALF" else
                    " ▷" if r.action in ("WIN","EW") and r.stake=="QUARTER" else
                    " ◑" if r.action=="CONSIDER" else " ✕" if r.action=="AVOID" else "  ")
            stake_suffix = f" [{r.stake}]" if r.action in ("WIN","EW") else ""
            action_str = f"{r.action}{stake_suffix}"
        lines.append(f"{flag} {i+1:>3}. {name_str:<23} {r.official_rating:>4}  {r.projected_rating:>5.1f}  {gap_str:>5}  {r.fractional_odds:>7}  {ev_str:>5}  {action_str}")
        if r.or_proxied: proxied_names.append((r.horse_name, r.proxy_note))
    lines.append(_divider(72))
    selections = [r for r in ranked if r.action in ("WIN","EW")]
    considers  = [r for r in ranked if r.action == "CONSIDER"]
    mode = getattr(ev_summary, 'selection_mode', 'SINGLE')
    coll_ev = getattr(ev_summary, 'basket_collective_ev', 0.0)
    if selections:
        lines.append("")
        if mode == "BASKET":
            lines.append(f"  BASKET ({len(selections)} selections  |  collective EV {coll_ev:.2f}x  |  level stakes profitable)")
        else:
            lines.append("  SELECTION")
        lines.append("  " + "─"*60)
        for r in selections:
            k_str = f"{r.kelly_pct:.2f}%" if hasattr(r,'kelly_pct') else ""
            v_str = f"{r.value_ratio:.2f}x" if hasattr(r,'value_ratio') else ""
            m_str = r.model_price_frac if hasattr(r,'model_price_frac') else ""
            lines.append(f"  ▶ {r.horse_name:<22} mkt {r.fractional_odds:<7}  model {m_str:<7}  value {v_str}  Kelly {k_str}  {r.action} [{r.stake}]")
    else:
        lines.append(""); lines.append("  NO SELECTION — insufficient value in this race")
    if considers:
        lines.append(""); lines.append("  ◑ CONSIDER — value present but below single selection threshold:")
        for r in considers:
            m_str = r.model_price_frac if hasattr(r,'model_price_frac') else ""
            v_str = f"{r.value_ratio:.2f}x" if hasattr(r,'value_ratio') else ""
            lines.append(f"    {r.horse_name:<22} mkt {r.fractional_odds:<7}  model {m_str:<7}  value {v_str}  edge {r.true_edge:+.1f}lb")
    if proxied_names:
        lines.append(""); lines.append("  * OR estimated from RPR — treat edge figures with caution:")
        for name, note in proxied_names: lines.append(f"    {name}: {note}")
    return "\n".join(lines)

_ANCHOR_RUN_TESTS = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 6 — Racing API Ingestor
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import requests, re, os, sys, time
from datetime import date, datetime

API_USER = os.environ.get('RACING_API_USER', '')
API_PASS = os.environ.get('RACING_API_PASS', '')
BASE_URL = 'https://api.theracingapi.com/v1'
RATE_LIMIT_DELAY = 0.22
EN_DASH = '\u2013'

class RacingAPIClient:
    def __init__(self, username, password):
        self.session = requests.Session()
        self.session.auth = (username, password)

    def _get(self, path, params=None):
        r = self.session.get(f'{BASE_URL}{path}', params=params, timeout=30)
        r.raise_for_status()
        return r.json()

    def get_racecards(self, day='today', course_ids=None, region_codes=None):
        params = {'day': day, 'limit': 500}
        if course_ids:   params['course_ids']   = course_ids
        if region_codes: params['region_codes'] = region_codes
        return self._get('/racecards/standard', params=params).get('racecards', [])

    def get_courses(self, region_codes=None):
        params = {}
        if region_codes: params['region_codes'] = region_codes
        return self._get('/courses', params=params).get('courses', [])

    def get_horse_results(self, horse_id, limit=15):
        try:
            time.sleep(RATE_LIMIT_DELAY)
            return self._get(f'/racecards/{horse_id}/results', params={'limit': limit}).get('results', [])
        except requests.HTTPError as e:
            code = e.response.status_code if e.response else '?'
            if code != 404: print(f'    warning: results HTTP {code} [{horse_id}]')
            return []
        except Exception as e:
            print(f'    warning: results error [{horse_id}]: {e}')
            return []

def _clean_rating(val):
    if val is None: return None
    s = str(val).strip()
    if s in (EN_DASH, '-', '', '0'): return None
    try: return int(float(s))
    except: return None

def _clean_dist_f(val):
    if val is None: return 10.0
    s = str(val).strip().lower().replace('f', '')
    try:
        f = float(s)
        if f > 0: return round(f, 2)
    except: pass
    miles = furls = yards = 0.0
    m = re.search(r'(\d+)m', str(val))
    if m: miles = float(m.group(1))
    fm = re.search(r'(\d+\.?\d*)f', str(val))
    if fm: furls = float(fm.group(1))
    ym = re.search(r'(\d+)y', str(val))
    if ym: yards = float(ym.group(1))
    total = miles*8 + furls + yards/220.0
    return round(total, 2) if total > 0 else 10.0

def _parse_position(pos):
    if pos is None: return None, None
    s = str(pos).strip().upper()
    NON = {'PU','P','F','FELL','UR','UNSEATED','BD','BROUGHT DOWN','RO','REF','REFUSED','SU','CO','DSQ','DQ','-',''}
    if s in NON:
        return None, ('PU' if s == 'P' else (s if s not in ('-','') else None))
    try: return int(float(s)), None
    except: return None, s

def _parse_btn(btn_val, position):
    if position == 1: return 0.0
    if btn_val is None: return None
    s = str(btn_val).strip()
    if s in ('-', EN_DASH, ''): return None
    try: return float(s)
    except: return None

def _norm_going(going_str):
    if not going_str: return 'GOOD'
    g = going_str.upper().strip()
    TABLE = {
        'HEAVY':'HEAVY', 'SOFT TO HEAVY':'HEAVY', 'VERY SOFT':'HEAVY',
        'SOFT':'SOFT', 'GOOD TO SOFT':'GOOD_TO_SOFT',
        'GOOD TO YIELDING':'GOOD_TO_SOFT', 'YIELDING TO SOFT':'GOOD_TO_SOFT',
        'YIELDING':'GOOD_TO_SOFT', 'GOOD':'GOOD', 'GOOD TO FIRM':'GOOD_TO_FIRM',
        'FIRM':'FIRM', 'HARD':'FIRM', 'FAST':'GOOD_TO_FIRM',
        'STANDARD TO FAST':'STANDARD', 'STANDARD':'STANDARD',
        'STANDARD TO SLOW':'GOOD_TO_SOFT', 'SLOW':'SOFT', 'MUDDY':'SOFT',
        'SLOPPY':'GOOD_TO_SOFT', 'HOLDING':'HEAVY',
    }
    for k, v in TABLE.items():
        if k in g: return v
    return 'GOOD'

def _extract_grade(race):
    pattern    = str(race.get('pattern',    '') or '').upper()
    race_class = str(race.get('class', race.get('race_class', '')) or '').upper()
    race_type  = str(race.get('type',  '') or '').upper()
    title      = str(race.get('race_name', '') or '').upper()
    if 'GRADE 1' in pattern or 'GROUP 1' in pattern: return 'G1'
    if 'GRADE 2' in pattern or 'GROUP 2' in pattern: return 'G2'
    if 'GRADE 3' in pattern or 'GROUP 3' in pattern: return 'G3'
    if 'LISTED'  in pattern: return 'LST'
    if 'HANDICAP' in title or 'HANDICAP' in race_type: return 'HCAP'
    if 'NOVICE' in title or 'NOVICES' in title: return 'NOV'
    for n in [1,2,3,4,5]:
        if f'CLASS {n}' in race_class or f'CLASS_{n}' in race_class: return f'CLS{n}'
    if 'MAIDEN' in title: return 'MDN'
    return 'DEFAULT'

def _course_code(name):
    TABLE = {
        'cheltenham':'CHE','ascot':'ASC','newmarket':'NMK','sandown':'SAN',
        'kempton':'KEM','goodwood':'GOO','york':'YOR','epsom':'EPS',
        'haydock':'HAY','newbury':'NWB','doncaster':'DON','leicester':'LEI',
        'nottingham':'NOT','windsor':'WIN','lingfield':'LIN',
        'wolverhampton':'WOL','southwell':'SOW','catterick':'CAT',
        'carlisle':'CAR','thirsk':'THI','beverley':'BEV','ripon':'RIP',
        'pontefract':'PON','redcar':'RED','musselburgh':'MUS',
        'hamilton':'HAM','ayr':'AYR','perth':'PER','stratford':'STR',
        'ludlow':'LUD','hereford':'HER','worcester':'WOR','warwick':'WAR',
        'huntingdon':'HUN','market rasen':'MRS','fakenham':'FAK',
        'exeter':'EXE','taunton':'TAU','wincanton':'WNC','plumpton':'PLU',
        'fontwell':'FON','brighton':'BTN','uttoxeter':'UTT','cartmel':'CTM',
        'bangor':'BAN','ffos las':'FFS','chepstow':'CPW',
        'chelmsford':'CHF','chelmsford (aw)':'CHF',
        'southwell (aw)':'SOW','wolverhampton (aw)':'WOL',
        'lingfield (aw)':'LIN','kempton (aw)':'KEM',
        'leopardstown':'LEO','fairyhouse':'FAI','punchestown':'PUN',
        'galway':'GAL','curragh':'CUR','naas':'NAA','navan':'NAV',
        'cork':'COR','gowran park':'GOW','tipperary':'TIP',
        'limerick':'LIM','dundalk':'DUN','down royal':'DWN',
        'killarney':'KIL','sligo':'SLI','roscommon':'ROS','tramore':'TRA',
        'clonmel':'CLO','ballinrobe':'BAL','bellewstown':'BEL',
        'auteuil':'AUT','pau':'PAU','compiegne':'COM','enghien':'ENG',
        'newcastle':'NEW','newcastle (aw)':'NEW',
    }
    key = name.lower().split('(')[0].strip()
    return TABLE.get(key, key[:3].upper())

def _irc_from_comment(comment):
    if not comment: return 'UNKNOWN'
    c = comment.lower()
    # FRONT_RUNNER — early-position phrases only
    if any(w in c for w in ['made all','led all','led from','led throughout','set the pace',
            'led 2f','led 3f','led 4f','in front','led early','jumped well in front','went clear']):
        return 'FRONT_RUNNER'
    if c.startswith('led') or ', led' in c[:30]: return 'FRONT_RUNNER'
    # PROMINENT — only with an explicit leader/pace reference
    if any(w in c for w in ['prominent','chased leader','chased leaders','close up',
            'tracked leader','tracked leaders','pressed leader','disputed lead',
            'disputed','handy']): return 'PROMINENT'
    # HOLD_UP — rear/held-up EARLY-POSITION phrases only.
    # NOTE: 'headway','progress','stayed on','late headway' REMOVED —
    # these describe finishing effort, not early position. A front-runner
    # that "made all, stayed on well" must NOT be pulled into HOLD_UP.
    if any(w in c for w in ['held up','held-up','towards rear','rear of field','back of field',
            'towards back','in rear','rearward','settled last','settled towards rear',
            'patiently ridden','waited with']):
        return 'HOLD_UP'
    if any(w in c for w in ['midfield','mid-field','in touch','middle of field']): return 'MIDFIELD'
    return 'UNKNOWN'

def _infer_running_style(position, margin, field_size):
    if position is None: return 'UNKNOWN'
    fs = max(field_size or 8, 4); pct = position / fs
    if position == 1 and margin is not None and margin >= 4.0: return 'FRONT_RUNNER'
    if pct <= 0.20: return 'PROMINENT'
    if pct >= 0.70: return 'HOLD_UP'
    return 'MIDFIELD'

def _decimal_to_fractional(dec):
    COMMON = {
        1.25:'1/4',1.33:'1/3',1.40:'2/5',1.50:'1/2',1.67:'4/6',1.80:'4/5',2.00:'EVS',
        2.25:'5/4',2.50:'6/4',2.75:'7/4',3.00:'2/1',3.50:'5/2',4.00:'3/1',4.50:'7/2',
        5.00:'4/1',5.50:'9/2',6.00:'5/1',7.00:'6/1',8.00:'7/1',9.00:'8/1',10.00:'9/1',
        11.00:'10/1',13.00:'12/1',15.00:'14/1',16.00:'15/1',17.00:'16/1',21.00:'20/1',
        26.00:'25/1',34.00:'33/1',41.00:'40/1',51.00:'50/1',101.00:'100/1',
    }
    closest = min(COMMON.keys(), key=lambda k: abs(k - dec))
    if abs(closest - dec) < 0.08: return COMMON[closest]
    num = dec - 1.0
    for d in [1,2,3,4,5,6,7,8,10,12,14,16,20,25]:
        n = round(num * d)
        if n > 0 and abs(n/d - num) < 0.03: return f'{n}/{d}'
    return f'{max(1,round(num))}/1'

_GW = {'G1':1.0,'G2':0.92,'G3':0.85,'LST':0.80,'CLS1':0.78,'CLS2':0.72,'CLS3':0.65,
       'CLS4':0.55,'CLS5':0.45,'HCAP':0.65,'NOV':0.70,'MDN':0.50,'DEFAULT':0.55,'UNK':0.55}

# ── AW course set ─────────────────────────────────────────────
# PATCH 2: Forces going='STANDARD' on all-weather tracks regardless
# of what the Racing API returns.
AW_COURSES = {
    'Chelmsford', 'Chelmsford (AW)', 'Chelmsford City',
    'Lingfield',  'Lingfield (AW)',  'Lingfield Park',
    'Kempton',    'Kempton (AW)',    'Kempton Park',
    'Wolverhampton', 'Wolverhampton (AW)',
    'Newcastle',  'Newcastle (AW)',
    'Southwell',  'Southwell (AW)',
    'Dundalk',
}

def _build_form_run(result_race, our_horse_id):
    our_runner = next((r for r in result_race.get('runners', []) if r.get('horse_id') == our_horse_id), None)
    if not our_runner: return None
    try: run_date = datetime.strptime(result_race['date'][:10], '%Y-%m-%d').date()
    except: run_date = date.today()
    position, comp_code = _parse_position(our_runner.get('position'))
    margin = _parse_btn(our_runner.get('ovr_btn', our_runner.get('btn')), position)
    rpr_int = _clean_rating(our_runner.get('rpr'))
    grade = _extract_grade(result_race)
    going_cat = _norm_going(result_race.get('going', ''))
    crs_code = _course_code(result_race.get('course', ''))
    distance_f = _clean_dist_f(result_race.get('dist_f', result_race.get('distance_f', result_race.get('dist', ''))))
    field_size = len(result_race.get('runners', []))
    local_or_map = {}
    for r in result_race.get('runners', []):
        h = r.get('horse', '')
        val = _clean_rating(r.get('or')) or (int(_clean_rating(r.get('rpr')) * 0.95) if _clean_rating(r.get('rpr')) else None)
        if h and val: local_or_map[h] = val
    comment = our_runner.get('comment', '') or ''
    style_str = _irc_from_comment(comment)
    if style_str == 'UNKNOWN': style_str = _infer_running_style(position, margin, field_size)
    style_to_irc = {'FRONT_RUNNER':'made all','PROMINENT':'tracked leader','MIDFIELD':'midfield','HOLD_UP':'held up','UNKNOWN':''}
    irc = style_to_irc.get(style_str, ''); gw = _GW.get(grade, 0.55)
    try: signal = _build_signal_from_irc(irc, gw)
    except: signal = None
    try: second_or, race_or_anchor = _build_anchors(position, margin, rpr_int, None, None, local_or_map)
    except: second_or = None; race_or_anchor = rpr_int
    fr = FormRun(run_date=run_date, position=position, completion_code=comp_code,
                 margin_lengths=margin, second_or=second_or, race_or_anchor=race_or_anchor,
                 grade_code=grade, signal=signal, chain_depth=0, chain_source_or=None)
    fr.going_category = going_cat; fr.course_code = crs_code; fr.distance_f = distance_f
    return fr

def _build_field(racecard, client):
    runners = racecard.get('runners', []); horses = []; odds_map = {}
    for i, runner in enumerate(runners):
        name = runner.get('horse', f'Runner {i+1}')
        horse_id = runner.get('horse_id', '')
        cloth = runner.get('number', str(i+1))
        try: cloth_int = int(str(cloth))
        except: cloth_int = i + 1
        or_int = _clean_rating(runner.get('ofr')); proxy_note = None
        if not or_int:
            rpr = _clean_rating(runner.get('rpr'))
            if rpr: or_int = int(rpr * 0.95); proxy_note = f'RPR {rpr}x0.95->{or_int}'
            else: or_int = 120; proxy_note = 'fallback OR 120'
        raw_results = client.get_horse_results(horse_id, limit=15) if horse_id else []
        runs = []
        for res_race in raw_results:
            try:
                fr = _build_form_run(res_race, horse_id)
                if fr: runs.append(fr)
            except: pass
        runs.sort(key=lambda r: r.run_date, reverse=True)
        h = Horse(name=name, official_rating=or_int, cloth_number=cloth_int, runs=runs)
        h._proxy_note = proxy_note; horses.append(h)
        best_decimal = None
        for bk in runner.get('odds', []):
            frac = bk.get('fractional', ''); dec = bk.get('decimal', '')
            if frac in ('SP','',None) or dec in ('SP','',None): continue
            try:
                d = float(dec)
                if d > 1.0 and (best_decimal is None or d > best_decimal): best_decimal = d
            except (ValueError, TypeError): pass
        if best_decimal and best_decimal > 1.0: odds_map[name] = _decimal_to_fractional(best_decimal)
        flag = f'  [{proxy_note}]' if proxy_note else ''
        odds_flag = f'  {odds_map.get(name,"--")}'
        print(f'    {cloth_int:>2}. {name:<32} OR={or_int:<4} {len(runs)} run(s){flag}{odds_flag}')
    return horses, odds_map


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 7 — CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import datetime, pytz
_uk  = pytz.timezone('Europe/London')
_now = datetime.datetime.now(_uk)
RACE_DATE = 'tomorrow' if _now.hour >= 18 else 'today'
REGION    = ['gb', 'ire']
COURSES   = []
ODDS      = {}

GITHUB_TOKEN  = os.environ['GITHUB_TOKEN']
ANTHROPIC_KEY = os.environ['ANTHROPIC_API_KEY']

print(f'📅 RACE_DATE={RACE_DATE}  |  UK time: {_now.strftime("%H:%M")}')
print(f'🌍 Region: {REGION}')
print(f'🏇 Courses: {"ALL" if not COURSES else ", ".join(COURSES)}')
print(f'💰 Manual odds: {sum(len(v) for v in ODDS.values())} races')

# ─────────────────────────────────────────────────────────────
# Draw bias — defined here (Cell 7) so Cell 8 can use it.
# ─────────────────────────────────────────────────────────────

def _course_slug(course):
    slug = course.lower()
    slug = slug.replace(' (aw)', '-aw').replace('(aw)', '-aw')
    slug = slug.replace('(','').replace(')','')
    slug = slug.replace(' ','-').replace('/','-')
    slug = re.sub(r'-+', '-', slug).strip('-')
    return slug

# Draw bias adjustments in lbs of rating.
# Generated from 12 months of handicap data, controlled for horse quality.
# Generated: 2026-05-20  — RE-RUN THE NOTEBOOK QUARTERLY TO REFRESH
DRAW_BIAS = {
    'ayr': {
        'mile': {'L': -0.87, 'M': +0.27, 'H': +0.38, 'n': 42, 'sig': False},
    },
    'bath': {
        'sprint': {'L': +3.04, 'M': -1.10, 'H': -1.26, 'n': 46, 'sig': True},
        'mile': {'L': +0.86, 'M': +1.35, 'H': -1.60, 'n': 36, 'sig': False},
    },
    'beverley': {
        'mile': {'L': +0.86, 'M': +1.59, 'H': -1.90, 'n': 31, 'sig': False},
    },
    'brighton': {
        'mile': {'L': +0.88, 'M': +0.84, 'H': -1.30, 'n': 43, 'sig': False},
    },
    'carlisle': {
        'mile': {'L': +0.01, 'M': +1.52, 'H': -1.23, 'n': 31, 'sig': False},
    },
    'catterick': {
        'sprint': {'L': +3.17, 'M': -0.27, 'H': -2.07, 'n': 38, 'sig': True},
    },
    'chelmsford-aw': {
        'sprint': {'L': +1.86, 'M': -1.41, 'H': -0.22, 'n': 67, 'sig': False},
        'mile': {'L': +1.08, 'M': +1.46, 'H': -1.95, 'n': 75, 'sig': True},
        'middle': {'L': -0.80, 'M': -0.45, 'H': +0.92, 'n': 39, 'sig': False},
    },
    'chepstow': {
        'sprint': {'L': +1.14, 'M': -0.13, 'H': -0.56, 'n': 33, 'sig': False},
        'mile': {'L': -1.52, 'M': +2.13, 'H': -0.64, 'n': 30, 'sig': False},
    },
    'doncaster': {
        'sprint': {'L': -0.34, 'M': +0.55, 'H': -0.19, 'n': 39, 'sig': False},
        'mile': {'L': +2.04, 'M': -0.19, 'H': -1.36, 'n': 55, 'sig': True},
        'middle': {'L': -0.58, 'M': +1.55, 'H': -0.79, 'n': 33, 'sig': False},
    },
    'dundalk-aw': {
        'mile': {'L': -0.19, 'M': +1.27, 'H': -0.80, 'n': 49, 'sig': False},
        'middle': {'L': +1.41, 'M': +0.26, 'H': -1.01, 'n': 35, 'sig': False},
    },
    'gowran-park': {
        'mile': {'L': +3.14, 'M': -2.09, 'H': -0.66, 'n': 32, 'sig': True},
    },
    'hamilton': {
        'sprint': {'L': +0.95, 'M': -0.37, 'H': -0.36, 'n': 51, 'sig': False},
    },
    'haydock': {
        'mile': {'L': +2.74, 'M': +0.89, 'H': -2.48, 'n': 36, 'sig': True},
    },
    'kempton-aw': {
        'sprint': {'L': +3.10, 'M': -0.04, 'H': -2.41, 'n': 55, 'sig': True},
        'mile': {'L': +0.84, 'M': +0.64, 'H': -1.16, 'n': 132, 'sig': True},
        'middle': {'L': +1.23, 'M': -0.48, 'H': -0.51, 'n': 54, 'sig': False},
    },
    'leicester': {
        'mile': {'L': +3.67, 'M': -0.29, 'H': -2.24, 'n': 34, 'sig': True},
    },
    'lingfield-aw': {
        'sprint': {'L': -0.59, 'M': +0.41, 'H': +0.07, 'n': 96, 'sig': False},
        'mile': {'L': -0.66, 'M': +0.35, 'H': +0.20, 'n': 125, 'sig': False},
        'middle': {'L': -0.67, 'M': +1.10, 'H': -0.42, 'n': 76, 'sig': False},
    },
    'musselburgh': {
        'mile': {'L': +0.82, 'M': -0.59, 'H': -0.12, 'n': 43, 'sig': False},
    },
    'newcastle-aw': {
        'sprint': {'L': -0.28, 'M': -0.63, 'H': +0.74, 'n': 160, 'sig': False},
        'mile': {'L': -1.11, 'M': +1.01, 'H': -0.04, 'n': 163, 'sig': False},
        'middle': {'L': -2.06, 'M': -0.33, 'H': +1.68, 'n': 73, 'sig': True},
    },
    'newmarket': {
        'mile': {'L': -0.84, 'M': -2.68, 'H': +2.77, 'n': 32, 'sig': False},
    },
    'newmarket-july': {
        'mile': {'L': -0.92, 'M': +3.68, 'H': -2.01, 'n': 30, 'sig': False},
    },
    'nottingham': {
        'sprint': {'L': -2.26, 'M': +0.12, 'H': +1.57, 'n': 31, 'sig': False},
    },
    'pontefract': {
        'sprint': {'L': +1.50, 'M': +0.85, 'H': -1.75, 'n': 36, 'sig': False},
    },
    'redcar': {
        'mile': {'L': +1.19, 'M': -1.30, 'H': +0.21, 'n': 35, 'sig': False},
    },
    'ripon': {
        'sprint': {'L': -0.43, 'M': +1.62, 'H': -0.93, 'n': 32, 'sig': False},
    },
    'southwell-aw': {
        'sprint': {'L': -0.21, 'M': +0.12, 'H': +0.06, 'n': 129, 'sig': False},
        'mile': {'L': +0.15, 'M': +0.63, 'H': -0.65, 'n': 167, 'sig': False},
        'middle': {'L': -1.35, 'M': +1.11, 'H': +0.06, 'n': 53, 'sig': False},
        'staying': {'L': -0.96, 'M': +1.13, 'H': -0.28, 'n': 46, 'sig': False},
    },
    'thirsk': {
        'sprint': {'L': -2.83, 'M': +0.42, 'H': +1.73, 'n': 35, 'sig': True},
    },
    'windsor': {
        'sprint': {'L': +0.49, 'M': -1.27, 'H': +0.61, 'n': 53, 'sig': False},
        'middle': {'L': +0.39, 'M': -0.67, 'H': +0.21, 'n': 37, 'sig': False},
    },
    'wolverhampton-aw': {
        'sprint': {'L': +1.11, 'M': +0.37, 'H': -1.15, 'n': 161, 'sig': True},
        'mile': {'L': -0.23, 'M': +0.71, 'H': -0.43, 'n': 244, 'sig': False},
        'middle': {'L': +1.61, 'M': +0.18, 'H': -1.16, 'n': 47, 'sig': True},
        'staying': {'L': +0.76, 'M': +1.80, 'H': -2.10, 'n': 36, 'sig': False},
    },
    'yarmouth': {
        'sprint': {'L': +2.95, 'M': +0.30, 'H': -2.08, 'n': 42, 'sig': True},
        'mile': {'L': +1.92, 'M': -0.30, 'H': -0.95, 'n': 46, 'sig': False},
    },
}

def _draw_adjustment(course, dist_f, draw, field_size):
    """Return the lbs adjustment for a given draw, from DRAW_BIAS.
    Non-significant cells with |value| > 2 lbs are dropped as noise.
    Returns 0.0 if no data for this course/distance."""
    if not draw or not field_size or draw < 1 or field_size < 4:
        return 0.0
    slug = _course_slug(course)
    slug = slug.replace('-ire', '')
    if dist_f <= 6.5:
        band = 'sprint'
    elif dist_f <= 9.5:
        band = 'mile'
    elif dist_f <= 12.5:
        band = 'middle'
    else:
        band = 'staying'
    cell = DRAW_BIAS.get(slug, {}).get(band)
    if not cell:
        return 0.0
    if draw <= field_size / 3:
        val = cell['L']
    elif draw <= 2 * field_size / 3:
        val = cell['M']
    else:
        val = cell['H']
    if not cell['sig'] and abs(val) > 2.0:
        return 0.0
    return val

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 8 — Fetch racecards + run ANCHOR for all meetings
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import time, datetime
from datetime import date

if not getattr(calculate_ev, '_patched', False):
    _orig_ev = calculate_ev
    def calculate_ev(*args, **kwargs):
        kwargs.pop('race_date', None)
        return _orig_ev(*args, **kwargs)
    calculate_ev._patched = True

client = RacingAPIClient(API_USER, API_PASS)
all_races = client.get_racecards(day=RACE_DATE, region_codes=REGION)
all_courses = sorted(set(r.get('course', '') for r in all_races))
print(f'📡 {len(all_races)} races across {len(all_courses)} meetings:')
for c in all_courses:
    n = sum(1 for r in all_races if r.get('course') == c)
    print(f'   {c} ({n} races)')

if COURSES:
    target_courses = [c for c in all_courses if any(req.lower() in c.lower() for req in COURSES)]
else:
    target_courses = all_courses

print(f'\n🏇 Running ANCHOR on: {", ".join(target_courses)}\n')

ALL_RESULTS = {}
import pytz
_uk_now = datetime.datetime.now(pytz.timezone('Europe/London'))
today_dt = (_uk_now + datetime.timedelta(days=1)).date() if _uk_now.hour >= 18 else _uk_now.date()
date_display = today_dt.strftime('%-d %B %Y')

for course in target_courses:
    course_races = sorted([r for r in all_races if r.get('course') == course],
                          key=lambda r: r.get('off_time', '00:00'))
    course_odds = ODDS.get(course, {})
    print(f'\n{"═"*60}\n  {course}  ({len(course_races)} races)\n{"═"*60}')
    course_results = []

    for racecard in course_races:
        race_time = racecard.get('off_time', '??:??')
        race_name = racecard.get('race_name', 'Unknown')
        going_str = racecard.get('going', 'Good')
        dist_f = _clean_dist_f(racecard.get('distance_f', racecard.get('distance_round', '')))
        today_going = _norm_going(going_str)
        # PATCH 3: force Standard going for AW tracks
        if course in AW_COURSES:
            today_going = 'STANDARD'
        race_grade = _extract_grade(racecard)
        crs_code = _course_code(course)
        n_runners = len(racecard.get('runners', []))
        print(f'\n  {race_time}  {race_name[:55]}')
        print(f'  {racecard.get("distance_round","?")} ({dist_f}f)  {going_str}  {race_grade}  {n_runners} runners')

        runner_extras = {}
        for runner in racecard.get('runners', []):
            name = runner.get('horse', '')
            if not name: continue
            t14 = runner.get('trainer_14_days', {}) or {}
            t14_runs = t14.get('runs', 0) or 0; t14_wins = t14.get('wins', 0) or 0
            t14_str = '{}/{}'.format(t14_wins, t14_runs) if t14_runs else ''
            runner_extras[name] = {
                'jockey':runner.get('jockey',''),'trainer':runner.get('trainer',''),
                'trainer_rtf':runner.get('trainer_rtf',''),'trainer_14':t14_str,
                'last_run':runner.get('last_run',None),'age':runner.get('age',''),
                'sex':runner.get('sex',''),'form':runner.get('form',''),
                'lbs':runner.get('lbs',''),'headgear':runner.get('headgear','') or '',
                'rpr':runner.get('rpr',None),'ts':runner.get('ts',None),
                'comment':runner.get('comment','') or '','spotlight':runner.get('spotlight','') or '',
            }
            going_parts = []
            for rc in course_races:
                t = rc.get('off_time',''); g = rc.get('going',''); d = rc.get('distance_round','')
                if g: going_parts.append('{} {}  {}'.format(t, d, g))
            course_going_report = '{}: {}'.format(course, ' | '.join(going_parts)) if going_parts else '{}: Going unknown'.format(course)

        try: horses, odds_map = _build_field(racecard, client)
        except Exception as e: print(f'  ❌ Field build error: {e}'); continue
        if not horses: print('  ⚠  No horses loaded — skipping'); continue

        race_odds = dict(odds_map)
        if race_time in course_odds:
            race_odds.update(course_odds[race_time])
            print(f'  ✚ Manual odds for {len(course_odds[race_time])} runners')

        try:
            avg_or, sd = field_stats(horses)
            rr_map = {h.name: calculate_private_rating(h, avg_or, sd, today_dt) for h in horses}
            pref_map = {h.name: calculate_preferences(h, rr_map[h.name].run_ratings, today_going, crs_code, dist_f) for h in horses}
            pace = classify_pace(horses, {h.name: rr_map[h.name].run_ratings for h in horses})
            damp_map = dampen_field(horses=horses, rating_results_map=rr_map, pref_results_map=pref_map,
                                    pace_result=pace, field_avg_or=avg_or,
                                    step_up_trip_map={}, step_up_class_map={})
 
            # ── Draw bias adjustment ──────────────────────────────
            # Shifts projected_rating by the quality-controlled draw
            # adjustment. Silent — not surfaced in UI. Feeds the EV calc.
            _field_size = len(horses)
            for _h in horses:
                if _h.name not in damp_map:
                    continue
                _adj = _draw_adjustment(course, dist_f, _h.cloth_number, _field_size)
                if _adj != 0.0:
                    _dr = damp_map[_h.name]
                    _dr.projected_rating = round(_dr.projected_rating + _adj, 2)
                    _dr.band_low  = round(_dr.band_low  + _adj, 2)
                    _dr.band_high = round(_dr.band_high + _adj, 2)
                    print(f'    draw-adj: {_h.name[:24]:<24} draw={_h.cloth_number} {_adj:+.2f}lb')
            # ──────────────────────────────────────────────────────
 
            ev_summary = calculate_ev(dampening_results=damp_map, odds_map=race_odds, race_name=race_name,
                                      field_avg_or=avg_or, pace_result=pace,
                                      race_grade=race_grade, distance_f=dist_f)
        except Exception as e:
            print(f'  ❌ Engine error: {e}')
            import traceback; traceback.print_exc()
            continue

        for r in ev_summary.results:
            if r.projected_rating is not None and r.projected_rating < 0:
                r.projected_rating = 0.0; r.band_low = 0.0; r.band_high = max(r.band_high, 0.0)

        meta = {
            'time':race_time, 'name':race_name, 'grade':race_grade,
            'dist':racecard.get('distance_round', f'{dist_f}f'), 'dist_f':dist_f,
            'going':today_going, 'course':crs_code, 'date':date_display,
            'pace_result':pace, 'horses':horses, 'runner_extras':runner_extras,
            'going_report':course_going_report,
        }
        course_results.append((meta, ev_summary))

        mode = ev_summary.selection_mode
        sels = [r for r in ev_summary.results if r.action in ('WIN', 'EW')]
        if sels:
            for s in sels:
                print(f'  ✅ {mode} ▶ {s.horse_name}  {s.fractional_odds}  EV {s.value_ratio:.2f}x  Kelly {s.kelly_pct:.2f}%')
        else:
            print(f'  ✅ {mode}')
        time.sleep(0.1)

    ALL_RESULTS[course] = course_results
    print(f'\n  {len(course_results)}/{len(course_races)} races processed')
    time.sleep(0.3)

RESULTS = []
for course_results in ALL_RESULTS.values(): RESULTS.extend(course_results)
COURSE = ', '.join(ALL_RESULTS.keys()) if ALL_RESULTS else 'Today'
print(f'\n{"═"*60}\n  TOTAL: {len(RESULTS)} races across {len(ALL_RESULTS)} meetings\n{"═"*60}')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CELL 12 — Publish to PaceMap
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

import requests, base64, json, time, datetime, re

GITHUB_REPO = 'ppcjobber/clerk-authentication-starter'




_FLAT_BIAS = {
    'Chester':   {'sprint':('STRONG',['L'],72,8),'mile':('STRONG',['L'],58,14),'middle':('MODERATE',['L','M'],42,22),'staying':('MINIMAL',['M'],35,28)},
    'Ascot':     {'sprint':('MODERATE',['H'],28,48),'mile':('SLIGHT',['M'],34,32),'middle':('MINIMAL',['M'],33,33),'staying':('MINIMAL',['M'],33,33)},
    'Newmarket': {'sprint':('SLIGHT',['H'],32,42),'mile':('SLIGHT',['H','M'],34,38),'middle':('MINIMAL',['M'],33,33),'staying':('MINIMAL',['M'],33,33)},
    'Goodwood':  {'sprint':('MODERATE',['L'],52,22),'mile':('SLIGHT',['L','M'],42,28),'middle':('MINIMAL',['M'],35,30),'staying':('MINIMAL',['M'],33,33)},
    'Epsom':     {'sprint':('MODERATE',['L','M'],48,18),'mile':('SLIGHT',['L','M'],40,26),'middle':('MINIMAL',['M'],35,28),'staying':('MINIMAL',['M'],33,33)},
    'York':      {'sprint':('SLIGHT',['H','M'],30,40),'mile':('MINIMAL',['M'],33,36),'middle':('MINIMAL',['M'],33,33),'staying':('MINIMAL',['M'],33,33)},
    'Haydock':   {'sprint':('MODERATE',['L'],50,20),'mile':('SLIGHT',['L','M'],40,28),'middle':('MINIMAL',['M'],35,30),'staying':('MINIMAL',['M'],33,33)},
    'Doncaster': {'sprint':('SLIGHT',['L','M'],40,32),'mile':('MINIMAL',['M'],34,34),'middle':('MINIMAL',['M'],33,33),'staying':('MINIMAL',['M'],33,33)},
    'Sandown':   {'sprint':('MODERATE',['L'],50,22),'mile':('SLIGHT',['L','M'],40,28),'middle':('MINIMAL',['M'],35,30),'staying':('MINIMAL',['M'],33,33)},
}

def _dist_band(f):
    if f<=6.5: return 'sprint'
    if f<=9.5: return 'mile'
    if f<=12.5: return 'middle'
    return 'staying'

def _draw_adv(draw, field_size, course, dist_f):
    if not draw or not field_size: return 'NEUTRAL'
    bias = _FLAT_BIAS.get(course,{}).get(_dist_band(dist_f))
    if not bias: return 'NEUTRAL'
    mag, fav, _, _ = bias
    pos = 'L' if draw<=field_size/3 else ('H' if draw>2*field_size/3 else 'M')
    if mag in ('MINIMAL','SLIGHT','UNKNOWN'): return 'NEUTRAL'
    if pos in fav: return 'FAVOURED'
    if ('L' in fav and pos=='H') or ('H' in fav and pos=='L'): return 'AGAINST'
    return 'NEUTRAL'

def _draw_bias_summary(course, dist_f):
    bias = _FLAT_BIAS.get(course,{}).get(_dist_band(dist_f))
    if not bias: return None
    mag, fav, lo, hi = bias
    return {'magnitude':mag,'favoured':'+'.join(fav),'low_pct':lo,'high_pct':hi}

def _is_flat(meta):
    name = meta.get('name','').lower()
    for kw in ['hurdle','chase','bumper','hunter','national hunt']:
        if kw in name: return False
    for kw in ['stakes','maiden','nursery','apprentice','fillies','handicap']:
        if kw in name: return True
    return False

def _build_runner_data(meta, ev, style_summary=None):
    is_flat = _is_flat(meta); course = meta.get('course',''); dist_f = meta.get('dist_f',10.0)
    ranked = sorted(ev.results, key=lambda r: r.projected_rating or 0, reverse=True)
    extras_map = meta.get('runner_extras', {}); out = []
    for r in ranked:
        draw = getattr(r,'draw',None) or getattr(r,'cloth_number',None)
        name = r.horse_name; s = style_summary.get(name, {}) if style_summary else {}
        extra = extras_map.get(name, {})
        out.append({
            'name':name,'or':r.official_rating,'style_code':s.get('style_code','U'),
            'finish_type':s.get('finish_type','E'),'dist_code':s.get('dist_code','B'),
            'going_flag':s.get('going_flag','UNKNOWN'),'note':'','draw':draw,
            'draw_adv':_draw_adv(draw,ev.field_size,course,dist_f) if is_flat else None,
            'jockey':extra.get('jockey',''),'trainer':extra.get('trainer',''),
            'trainer_rtf':extra.get('trainer_rtf',''),'trainer_14':extra.get('trainer_14',''),
            'last_run':extra.get('last_run',None),'age':extra.get('age',''),
            'sex':extra.get('sex',''),'form':extra.get('form',''),
            'lbs':extra.get('lbs',''),'headgear':extra.get('headgear',''),
            'rpr':extra.get('rpr',None),'ts':extra.get('ts',None),
            'comment':extra.get('comment',''),'spotlight':extra.get('spotlight',''),
            '_projected':round(r.projected_rating,1) if r.projected_rating else 0.0,
            '_model_price':r.model_price_frac,
            '_mkt_price':r.fractional_odds if r.fractional_odds not in ('n/a','--','',None) else None,
            '_win_pct':round((r.model_prob or 0)*100,1),
            '_edge':round(r.true_edge,1) if isinstance(r.true_edge,(int,float)) else None,
            '_action':r.action,
        })
    return out

def _build_style_summary(horses, today_going, today_dist_f=None, is_aw=False):
    # PATCH 4: STANDARD added to GOING_SIMILAR
    GOING_SIMILAR = {
        'HEAVY':        {'HEAVY','SOFT'},
        'SOFT':         {'HEAVY','SOFT','GOOD_TO_SOFT'},
        'GOOD_TO_SOFT': {'SOFT','GOOD_TO_SOFT','GOOD'},
        'GOOD':         {'GOOD_TO_SOFT','GOOD','GOOD_TO_FIRM'},
        'GOOD_TO_FIRM': {'GOOD','GOOD_TO_FIRM','FIRM'},
        'FIRM':         {'GOOD_TO_FIRM','FIRM'},
        'STANDARD':     {'STANDARD'},
    }
    similar = GOING_SIMILAR.get(today_going, {today_going}); summary = {}
    for horse in horses:
        runs = horse.runs or []; name = horse.name; notes = []
        sc = {'L':0,'P':0,'M':0,'H':0}
        for run in runs:
            sig = getattr(run,'signal',None); rs = str(getattr(sig,'running_style','') or '').upper()
            if 'FRONT' in rs: sc['L'] += 1
            elif 'PROMINENT' in rs: sc['P'] += 1
            elif 'MIDFIELD' in rs: sc['M'] += 1
            elif 'HOLD' in rs: sc['H'] += 1
        total_styled = sum(sc.values())
        modal = max(('L','P','M','H'), key=lambda k: sc[k]) if total_styled else None
        modal_n = sc[modal] if modal else 0
        STYLE_LABELS = {'L':'front runner','P':'prominent','M':'midfield','H':'hold up'}
        if total_styled == 0:
            style_label = 'style unknown'
            if len(runs) < 3: notes.append('lightly raced')
        elif modal_n >= total_styled * 0.70: style_label = 'confirmed {}'.format(STYLE_LABELS[modal])
        elif modal_n >= total_styled * 0.45: style_label = 'usually {}'.format(STYLE_LABELS[modal])
        else:
            sorted_styles = sorted(sc.items(), key=lambda x: x[1], reverse=True)
            top2 = [STYLE_LABELS[k] for k,v in sorted_styles[:2] if v > 0]
            style_label = 'versatile ({})'.format('/'.join(top2)) if top2 else 'versatile'
        fq = {'S':0,'E':0,'F':0}
        completed = [r for r in runs if r.position is not None and r.position > 0]
        for run in completed:
            field = getattr(run, 'field_size', None) or 10; pos = run.position
            if pos == 1 or pos <= max(2, field * 0.25): fq['S'] += 1
            elif pos >= field * 0.75: fq['F'] += 1
            else: fq['E'] += 1
        total_fq = sum(fq.values())
        if total_fq == 0: finish_type = 'E'; finish_label = 'no completions'
        elif fq['S'] >= total_fq * 0.55: finish_type = 'S'; finish_label = 'strong finisher'
        elif fq['F'] >= total_fq * 0.55:
            finish_type = 'F'; finish_label = 'fading profile'; notes.append('fades in finish')
        else: finish_type = 'E'; finish_label = 'even pace'
        if not runs: going_flag = 'UNKNOWN'; going_str = 'no form data'
        elif is_aw:
            aw_runs = [r for r in runs if getattr(r, 'surface', '') in ('AW','Polytrack','Tapeta','Fibresand')]
            wins = sum(1 for r in runs if r.position == 1)
            placed = sum(1 for r in runs if r.position and r.position <= 3)
            if len(runs) >= 3 and not aw_runs:
                going_flag = 'CONCERN'; going_str = 'no AW form ({} turf runs)'.format(len(runs)); notes.append('NO AW FORM')
            else: going_flag = 'OK'; going_str = '{} runs — {}W {}P'.format(len(runs), wins, placed)
        else:
            wins = sum(1 for r in runs if r.position == 1)
            placed = sum(1 for r in runs if r.position and r.position <= 3)
            going_flag = 'OK'; going_str = '{} runs — {}W {}P'.format(len(runs), wins, placed)
        if today_dist_f is None or not runs: dist_code = 'UNKNOWN'
        else:
            dist_runs = [r for r in runs if r.distance_f and abs(r.distance_f - today_dist_f) <= 1.5]
            if not dist_runs:
                shorter = [r for r in runs if r.distance_f and r.distance_f < today_dist_f - 1.5]
                longer  = [r for r in runs if r.distance_f and r.distance_f > today_dist_f + 1.5]
                dist_code = 'B+' if (len(shorter) >= len(longer) and shorter) else ('B-' if longer else 'UNKNOWN')
            else:
                trip_placed = sum(1 for r in dist_runs if r.position and r.position <= 3)
                dist_code = 'B+' if (fq.get('S',0) > fq.get('F',0) and trip_placed >= 1) else 'B'
        summary[name] = {
            'style_code':modal or 'U','style_label':style_label,'finish_type':finish_type,
            'finish_label':finish_label,'going_flag':going_flag,'going_str':going_str,
            'dist_code':dist_code,'notes':notes,'n_runs':len(runs),'sc':sc,'fq':fq,
        }
    return summary

def _build_pace_dynamic(runners, style_summary, ev, meta):
    field_size = ev.field_size
    buckets = {'L':[],'P':[],'M':[],'H':[],'U':[]}
    for r in runners:
        name = r['name']; s = style_summary.get(name, {}) if style_summary else {}
        code = s.get('style_code','U'); or_ = r.get('or',0) or 0
        sc = s.get('sc',{'L':0,'P':0,'M':0,'H':0}); buckets[code].append((or_, name, sc))
    for k in buckets: buckets[k].sort(reverse=True)
    leader_names = [n for _,n,_ in buckets['L']]
    prominent_names = [n for _,n,_ in buckets['P']]
    midfield_names = [n for _,n,_ in buckets['M']]
    holdup_names = [n for _,n,_ in buckets['H']]
    unknown_names = [n for _,n,_ in buckets['U']]
    midfield_names.extend(unknown_names)

    # ── Redistribution tracking ───────────────────────────────
    # When a bucket holds more runners than the map can show, the
    # overflow is moved to midfield FOR DISPLAY ONLY. We record each
    # moved horse's TRUE style here so the narrative can be honest:
    # the horse is still really a prominent/hold-up type, we just
    # don't expect it to get that position in this particular race.
    redistributed = {}  # name -> {'true_style','mapped_to','reason'}

    max_prom = max(2, round(field_size * 0.40))
    if len(prominent_names) > max_prom:
        overflow = prominent_names[max_prom:]; prominent_names = prominent_names[:max_prom]
        midfield_names.extend(overflow)
        for n in overflow:
            redistributed[n] = {
                'true_style':'prominent', 'mapped_to':'midfield',
                'reason':'more prominent types than can hold that position in this field'}

    max_hold = max(2, round(field_size * 0.30))
    if len(holdup_names) > max_hold:
        holdup_sorted = sorted(buckets['H'], reverse=True)
        holdup_names = [n for _,n,_ in holdup_sorted[:max_hold]]
        overflow = [n for _,n,_ in holdup_sorted[max_hold:]]; midfield_names.extend(overflow)
        for n in overflow:
            redistributed[n] = {
                'true_style':'hold up', 'mapped_to':'midfield',
                'reason':'more hold-up types than the map shows at the rear'}

    if not leader_names:
        if prominent_names: promoted = prominent_names[0]; prominent_names = prominent_names[1:]; leader_names = [promoted]
        else:
            best_lead = None; best_lead_or = -1
            for r in runners:
                name = r['name']; s = style_summary.get(name, {}) if style_summary else {}
                sc = s.get('sc',{'L':0,'P':0,'M':0,'H':0}); or_ = r.get('or',0) or 0
                if sc.get('L',0) >= 1 and or_ > best_lead_or: best_lead = name; best_lead_or = or_
            if best_lead:
                for lst in [prominent_names, midfield_names, holdup_names]:
                    if best_lead in lst: lst.remove(best_lead)
                leader_names = [best_lead]
                # A horse with no confirmed lead history promoted to lead
                # purely because the race lacks one — flag it honestly.
                redistributed[best_lead] = {
                    'true_style':'prominent', 'mapped_to':'front',
                    'reason':'no confirmed front runner in the race, so the most likely pace-setter is shown on the lead'}

    n_leads = len(leader_names); n_prom = len(prominent_names)
    n_mid = len(midfield_names); n_hold = len(holdup_names)
    if n_leads == 0: dynamic = 'No confirmed front runner — field likely to drift into a sprint finish.'
    elif n_leads == 1: dynamic = 'One confirmed front runner ({}) — sustainable pace, not pressured.'.format(leader_names[0])
    elif n_leads == 2: dynamic = 'Two confirmed front runners ({}) — contested early, honest gallop likely.'.format(', '.join(leader_names[:2]))
    else: dynamic = '{} confirmed front runners ({}) — strong gallop almost guaranteed. Stamina tested.'.format(n_leads, ', '.join(leader_names[:3]))

    congestion = ''
    if n_prom >= field_size * 0.4 and n_prom >= 4:
        congestion = ' Prominent congestion: {} of {} runners want that position — traffic danger 2-3 out.'.format(n_prom, field_size)

    # ── Honest redistribution sentence in the pace dynamic ─────
    redis_note = ''
    if redistributed:
        moved_to_mid = [n for n, d in redistributed.items() if d['mapped_to'] == 'midfield']
        if moved_to_mid:
            shown = ', '.join(moved_to_mid[:3])
            more = '' if len(moved_to_mid) <= 3 else ' and others'
            redis_note = (' Note: this field contains more pace than the map can place — '
                          '{}{} are genuine prominent/hold-up types shown in midfield because '
                          'they are unlikely to get their natural position in this race shape.').format(shown, more)

    return dynamic + congestion + redis_note, {
        'n_leads':n_leads,'n_prom':n_prom,'n_mid':n_mid,'n_hold':n_hold,
        'leader_names':leader_names,'prominent_names':prominent_names,
        'midfield_names':midfield_names,'holdup_names':holdup_names,
        'redistributed':redistributed,
    }

def _call_claude(prompt, max_tokens=2800):
    for attempt in range(4):
        resp = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={'Content-Type':'application/json','x-api-key':ANTHROPIC_KEY,'anthropic-version':'2023-06-01'},
            json={'model':'claude-sonnet-4-20250514','max_tokens':max_tokens,'messages':[{'role':'user','content':prompt}]})
        data = resp.json()
        if 'error' in data:
            err_type = data['error'].get('type','') if isinstance(data['error'],dict) else str(data['error'])
            if 'overload' in str(err_type).lower() and attempt < 3:
                wait = 15 * (attempt + 1); print('         overloaded — waiting {}s...'.format(wait)); time.sleep(wait); continue
            raise ValueError(data['error'])
        return ''.join(b['text'] for b in data.get('content',[]) if b.get('type')=='text')
    raise ValueError('Claude overloaded after 4 attempts')

def _narrative_prompt(meta, ev, runners, going_report, is_flat,
                      style_summary=None, pace_dynamic_str=None, pace_info=None, horses_list=None):
    runners_sorted = sorted(runners, key=lambda r: r.get('_projected') or 0, reverse=True)
    # PATCH 4: STANDARD added to GOING_SIMILAR
    GOING_SIMILAR = {
        'HEAVY':        {'HEAVY','SOFT'},
        'SOFT':         {'HEAVY','SOFT','GOOD_TO_SOFT'},
        'GOOD_TO_SOFT': {'SOFT','GOOD_TO_SOFT','GOOD'},
        'GOOD':         {'GOOD_TO_SOFT','GOOD','GOOD_TO_FIRM'},
        'GOOD_TO_FIRM': {'GOOD','GOOD_TO_FIRM','FIRM'},
        'FIRM':         {'GOOD_TO_FIRM','FIRM'},
        'STANDARD':     {'STANDARD'},
    }
    today_going_key = meta.get('going','GOOD').upper().replace(' ','_')
    similar = GOING_SIMILAR.get(today_going_key, {today_going_key})
    runner_lines = []
    for r in runners_sorted:
        name = r['name']; s = style_summary.get(name, {}) if style_summary else {}
        notes_str = ' | '.join(s.get('notes',[])) if s.get('notes') else ''
        flag = ' \u26a0 {}'.format(notes_str) if notes_str else ''
        redis = (pace_info or {}).get('redistributed', {}).get(name)
        redis_str = ''
        if redis:
            redis_str = ' [REDISTRIBUTED: true {} \u2014 mapped to {} because {}]'.format(
                redis['true_style'], redis['mapped_to'], redis['reason'])
        draw_str = ''
        if is_flat and r.get('draw'): draw_str = ' draw={}({})'.format(r['draw'], r.get('draw_adv',''))
        jockey = r.get('jockey','') or ''; trainer = r.get('trainer','') or ''
        trainer_14 = r.get('trainer_14','') or ''; last_run = r.get('last_run',None)
        age = r.get('age','') or ''; form_str_rc = r.get('form','') or ''
        lbs = r.get('lbs','') or ''; headgear = r.get('headgear','') or ''
        rpr = r.get('rpr',None); ts = r.get('ts',None); spotlight = r.get('spotlight','') or ''
        extras_parts = []
        if jockey: extras_parts.append('J:{}'.format(jockey))
        if trainer: extras_parts.append('T:{}'.format(trainer))
        if trainer_14: extras_parts.append('T14:{}'.format(trainer_14))
        if last_run is not None: extras_parts.append('LR:{}d'.format(last_run))
        if age: extras_parts.append('age:{}'.format(age))
        if form_str_rc: extras_parts.append('form:{}'.format(form_str_rc))
        if lbs: extras_parts.append('{}lbs'.format(lbs))
        if headgear: extras_parts.append('hg:{}'.format(headgear))
        if rpr: extras_parts.append('RPR:{}'.format(rpr))
        if ts: extras_parts.append('TS:{}'.format(ts))
        extras_str = '  [{}]'.format(' | '.join(extras_parts)) if extras_parts else ''
        horse_obj = next((h for h in (horses_list or []) if h.name == name), None)
        hist_str = ''
        if horse_obj and horse_obj.runs:
            recent = sorted(horse_obj.runs, key=lambda x: x.run_date, reverse=True)[:5]
            pos_parts = []
            for run in recent:
                pos = str(run.position) if run.position else '?'
                dist = '{:.0f}f'.format(run.distance_f) if run.distance_f else '?f'
                going_cat = (getattr(run,'going_category','') or '')[:3]
                pos_parts.append('{}/{}/{}'.format(pos, dist, going_cat or '?'))
            hist_str = ' hist: {}'.format(' '.join(pos_parts))
            going_runs = [r2 for r2 in horse_obj.runs if getattr(r2,'going_category','') in similar]
            if going_runs:
                gw = sum(1 for r2 in going_runs if r2.position == 1)
                gp = sum(1 for r2 in going_runs if r2.position and r2.position <= 3)
                hist_str += ' sim-going:{}/{} {}W{}P'.format(len(going_runs), len(horse_obj.runs), gw, gp)
        spot_str = '\n    SP: {}'.format(spotlight[:160]) if spotlight else ''
        runner_lines.append('  {:<26} OR={:<4} style={} going={}{}{}{}{}{}{}'.format(
            name, r.get('or','--'), s.get('style_code','U'), s.get('going_flag','?'),
            flag, redis_str, draw_str, extras_str, hist_str, spot_str))
    runners_block = '\n'.join(runner_lines)
    draw_note = ''
    if is_flat:
        bias = _draw_bias_summary(meta.get('course',''), meta.get('dist_f',10.0))
        if bias: draw_note = '\nDRAW: {} — {} draws favoured.'.format(bias['magnitude'], bias['favoured'])
    going_display = meta.get('going','').replace('_',' ')
    race_type = 'FLAT' if is_flat else 'JUMPS'
    course_name = meta.get('course','')
    is_aw = any(x in course_name.upper() for x in ['(AW)','ALL WEATHER','ALL-WEATHER','DUNDALK','LINGFIELD','KEMPTON','WOLVERHAMPTON','CHELMSFORD','NEWCASTLE'])
    surface = 'ALL-WEATHER (Polytrack/Tapeta)' if is_aw else 'TURF'
    return (
        'You are PaceMap — a race shape analysis tool. Produce a structured analysis for this {} race.\n'
        'Output EXACTLY the delimited format below — no extra text before or after.\n\n'
        'RACE: {} {}\n'
        'COURSE: {} | DIST: {} | GOING: {} | SURFACE: {} | RUNNERS: {}{}\n'
        'GOING REPORT: {}\n\n'
        'PACE DYNAMIC: {}\n\n'
        'RUNNERS (style: L=Lead P=Prominent M=Midfield H=HoldUp | going: OK/CONCERN/UNKNOWN\n'
        ' extras: J=jockey T=trainer T14=trainer 14-day W/R LR=days since last run RPR=Racing Post Rating TS=TopSpeed\n'
        ' form=recent form string e.g. 1-2-3 | hist=last 5 runs pos/dist/going | SP=spotlight comment):\n'
        '{}\n\n'
        'Output this EXACT format with these EXACT delimiters:\n\n'
        '%%PACE_DYNAMIC%%\n'
        '[3-4 sentences. Name specific horses. Describe how this race will be run from flag to finish. '
        'Name the pace-setters and what they create. Conclude with a clear tactical verdict: '
        'which running style profile this race shape favours and why. '
        'If any runners are tagged [REDISTRIBUTED] in the runner list, add one sentence stating that '
        'the field holds more front-running or prominent types than can all get their position, and name '
        'which are most likely to actually secure it. '
        'Be direct. No hedging.]\n\n'
        '%%SCENARIOS%%\n'
        'SCENARIO_A|[3-5 word title]|[probability 0-100]|[one-line trigger starting with "If..."]|'
        '[2 sentences: how it unfolds and why this profile benefits]|[winner1,winner2]|[other1,other2]\n'
        'SCENARIO_B|[3-5 word title]|[probability 0-100]|[one-line trigger]|[2 sentences]|[winner1,winner2]|[other1,other2]\n'
        'SCENARIO_C|[3-5 word title]|[probability 0-100]|[one-line trigger]|[2 sentences]|[winner1,winner2]|[other1,other2]\n'
        'SCENARIO_D|[3-5 word title]|[probability 0-100]|[one-line trigger]|[2 sentences]|[winner1,winner2]|[other1,other2]\n\n'
        '%%NOTES%%\n'
        '[horse_name]|[one concise sentence covering: running style, jockey/trainer angle if relevant, '
        'going or distance concern, headgear, days since last run if notable, or key tactical factor]\n'
        '...one line per runner, all runners\n'
        'HONESTY RULE: If a runner is tagged [REDISTRIBUTED] in the runner list, its note MUST state the '
        'horse\'s true running style first, then explain the shown map position as a consequence of the '
        'race shape — e.g. "a natural front-runner, but with several confirmed leaders here is likely to be '
        'taken back into midfield". NEVER describe a redistributed horse as though the shown position were '
        'its natural style.\n\n'
        '%%WATCH%%\n'
        '[severity: danger/warn/info]|[one concrete sentence naming specific horse(s) and what to watch]\n'
        '...3-5 watch points total\n\n'
        'PROBABILITY RULES: Probabilities across A+B+C+D must sum to exactly 100. '
        'DO NOT distribute evenly. Base probabilities solely on the evidence. '
        'If evidence strongly favours one scenario, assign it 55-70%. '
        'Example strong lean: 60/20/12/8. Example moderate lean: 45/30/15/10.\n\n'
        'OTHER RULES: Name horses throughout. No generic phrases. No tips or selections.\n'
        'Winners are horses whose style/profile SUITS that scenario.\n'
        'The %%PACE_DYNAMIC%% must end with a clear directional verdict on which profile benefits.'
    ).format(race_type, meta['time'], meta['name'], meta.get('course',''), meta['dist'],
             going_display, surface, ev.field_size, draw_note, going_report,
             pace_dynamic_str or 'No pace dynamic computed.', runners_block)

def _parse_narrative(raw, runners):
    def _section(text, marker):
        pat = r'%%{}%%(.*?)(?=%%[A-Z_]+%%|$)'.format(marker)
        m = re.search(pat, text, re.DOTALL)
        return m.group(1).strip() if m else ''
    pace_dynamic = _section(raw, 'PACE_DYNAMIC')
    scenarios = []
    scen_block = _section(raw, 'SCENARIOS')
    for letter in ['A','B','C','D']:
        pat = r'SCENARIO_{}[|\s](.+?)(?=SCENARIO_[ABCD]|$)'.format(letter)
        m = re.search(pat, scen_block, re.DOTALL)
        if not m: continue
        parts = [p.strip() for p in m.group(1).split('|')]
        if len(parts) < 6: continue
        try: prob = int(re.sub(r'[^0-9]','',parts[1]))
        except: prob = 25
        winners = [h.strip() for h in parts[4].split(',') if h.strip()]
        others  = [h.strip() for h in parts[5].split(',') if h.strip()]
        scenarios.append({'label':letter,'title':parts[0],'prob':prob,'trigger':parts[2],'body':parts[3],'winners':winners,'others':others})
    notes_map = {}
    for line in _section(raw, 'NOTES').split('\n'):
        line = line.strip()
        if '|' not in line: continue
        idx = line.index('|'); horse = line[:idx].strip(); note = line[idx+1:].strip()
        if horse and note: notes_map[horse] = note
    watch_points = []
    for line in _section(raw, 'WATCH').split('\n'):
        line = line.strip()
        if '|' not in line: continue
        idx = line.index('|'); sev = line[:idx].strip().lower(); text = line[idx+1:].strip()
        if sev not in {'danger','warn','info'}: sev = 'info'
        if text: watch_points.append({'severity':sev,'text':text})
    return pace_dynamic, scenarios, notes_map, watch_points
  
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Consistency guardrail — catches AI notes that contradict the
# engine's own style classification before they reach a customer.
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_HOLDUP_LANGUAGE = [
    'held up', 'held-up', 'hold up', 'hold-up', 'from the rear', 'off the pace',
    'towards the rear', 'at the back', 'settled last', 'patiently ridden',
    'waited with', 'dropped out', 'in rear', 'rearward', 'last of all',
]
_LEAD_LANGUAGE = [
    'make all', 'makes all', 'made all', 'set the pace', 'sets the pace',
    'from the front', 'leads the field', 'go on early', 'bowl along in front',
    'dictate from the front',
]

def _consistency_check(races_data):
    """Flag any runner whose AI note contradicts the engine's style code.
    Returns a list of conflict dicts. Does not modify races_data."""
    conflicts = []
    for race in races_data:
        if race.get('skipped'):
            continue
        engine_style = {}
        for n in race.get('leads', []):     engine_style[n] = 'L'
        for n in race.get('prominent', []): engine_style.setdefault(n, 'P')
        for n in race.get('midfield', []):  engine_style.setdefault(n, 'M')
        for n in race.get('holdup', []):    engine_style.setdefault(n, 'H')
        for r in race.get('runners_data', []):
            name = r.get('name', ''); note = (r.get('note', '') or '').lower()
            if not note:
                continue
            code = engine_style.get(name, r.get('style_code', 'U'))
            if code == 'L' and any(p in note for p in _HOLDUP_LANGUAGE):
                conflicts.append({'race':race['time'],'horse':name,
                                  'engine':'front-runner','note_says':'hold-up','note':r.get('note','')})
            elif code == 'H' and any(p in note for p in _LEAD_LANGUAGE):
                conflicts.append({'race':race['time'],'horse':name,
                                  'engine':'hold-up','note_says':'front-runner','note':r.get('note','')})
    return conflicts

def _gh_push(path, content_str, message):
    url = 'https://api.github.com/repos/{}/contents/{}'.format(GITHUB_REPO, path)
    headers = {'Authorization':'token {}'.format(GITHUB_TOKEN),'Accept':'application/vnd.github.v3+json'}
    r = requests.get(url, headers=headers)
    sha = r.json().get('sha') if r.status_code == 200 else None
    payload = {'message':message,'content':base64.b64encode(content_str.encode()).decode()}
    if sha: payload['sha'] = sha
    r = requests.put(url, headers=headers, json=payload)
    return r.status_code in (200,201)

def _update_archive(slug, course, date_str, going, n_races):
    url = 'https://api.github.com/repos/{}/contents/app/archive/page.tsx'.format(GITHUB_REPO)
    headers = {'Authorization':'token {}'.format(GITHUB_TOKEN),'Accept':'application/vnd.github.v3+json'}
    r = requests.get(url, headers=headers)
    if r.status_code != 200: print('  warning: archive fetch {}'.format(r.status_code)); return False
    current = base64.b64decode(r.json()['content']).decode(); sha = r.json()['sha']
    if '"{}"'.format(slug) in current: print('  Archive entry already exists — skipping duplicate'); return True
    course_slug = _course_slug(course)
    entry = ('\n  {{\n    slug:       "{}",\n    date:       "{}",\n    label:      "{} \u2014 {}",\n'
             '    going:      "{}",\n    races:      {},\n    courseSlug: "{}",\n    latest:     true,\n  }},').format(
        slug, date_str, course, date_str, going, n_races, course_slug)
    current = current.replace('    latest:     true,', '    latest:     false,')
    current = current.replace('    latest: true,', '    latest: false,')
    current = re.sub(r'const MEETINGS: any\[\] = \[', 'const MEETINGS: any[] = [' + entry, current)
    payload = {'message':'archive: add {}'.format(slug),'content':base64.b64encode(current.encode()).decode(),'sha':sha}
    r = requests.put(url, headers=headers, json=payload)
    return r.status_code in (200, 201)

import json as _json
import os as _os
import pytz as _pytz
from datetime import datetime as _datetime, timedelta as _timedelta

_UK_TZ = _pytz.timezone('Europe/London')
TWEET_QUEUE_FILE = 'tweets_queue.json'

def _json_exists(path):
    return _os.path.exists(path)

def _parse_race_time(time_str, race_date_str):
    for fmt in ['%-d %B %Y %H:%M', '%d %B %Y %H:%M']:
        try:
            dt = _datetime.strptime('{} {}'.format(race_date_str, time_str), fmt)
            return _UK_TZ.localize(dt)
        except Exception:
            continue
    return None

# ── PATCH 5a: Tweet email function ────────────────────────────
def _send_tweet_email(queue, course, race_date_str):
    """
    Email tweet copy to ppcjobber@gmail.com after each pipeline run.
    Requires GitHub Actions secrets: GMAIL_APP_PASSWORD, GMAIL_FROM.
    Create an app password at: myaccount.google.com/apppasswords
    """
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    gmail_pass = os.environ.get('GMAIL_APP_PASSWORD', '')
    gmail_from = os.environ.get('GMAIL_FROM', 'ppcjobber@gmail.com')

    if not gmail_pass:
        print('  GMAIL_APP_PASSWORD not set -- skipping tweet email')
        return
    if not queue:
        print('  Tweet queue empty -- skipping email')
        return

    subject = 'PaceMap tweets -- {} {}'.format(course, race_date_str)
    lines = [
        'PaceMap tweet copy -- {} {}'.format(course, race_date_str),
        '=' * 60, '',
        'Copy each tweet into Buffer manually.', '',
    ]
    for i, tweet in enumerate(queue, 1):
        lines.append('TWEET {} [{}]'.format(i, tweet.get('type','')))
        lines.append('Scheduled: {}'.format(tweet.get('scheduled_time','')))
        lines.append('-' * 40)
        lines.append(tweet.get('text',''))
        lines.append('')

    body = '\n'.join(lines)
    msg = MIMEMultipart()
    msg['From']    = gmail_from
    msg['To']      = 'ppcjobber@gmail.com'
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(gmail_from, gmail_pass)
            server.send_message(msg)
        print('  Tweet email sent -- {} tweets'.format(len(queue)))
    except Exception as e:
        print('  Tweet email failed: {}'.format(e))


def _write_tweet_queue(slug, course, race_date_str, races_data):
    queue = []
    if _json_exists(TWEET_QUEUE_FILE):
        with open(TWEET_QUEUE_FILE) as f:
            try: queue = _json.load(f)
            except Exception: queue = []
    url = 'pacemap.co.uk/meetings/{}'.format(slug)
    for race in races_data:
        if race.get('skipped'): continue
        race_time = race.get('time',''); dist = race.get('dist','')
        going = race.get('going',''); n_runners = race.get('runners',0)
        dynamic = race.get('paceDynamic',''); scenarios = race.get('scenarios',[])
        if not dynamic or not scenarios: continue
        off_dt = _parse_race_time(race_time, race_date_str)
        if not off_dt: continue
        short_dynamic = (dynamic[:240] + '...') if len(dynamic) > 240 else dynamic
        tweet1_text = '{} {} \u00b7 {} \u00b7 {} \u00b7 {} runners\n\n{}\n\n{}'.format(
            course, race_time, dist, going, n_runners, short_dynamic, url)
        queue.append({'scheduled_time':(off_dt - _timedelta(minutes=30)).isoformat(),
                      'text':tweet1_text,'type':'race_dynamic','course':course,'race_time':race_time})
        sc_lines = '\n'.join('{}) {} ({}%) — {} benefit'.format(
            s['label'], s['title'], s['prob'], ', '.join(s['winners'][:2])) for s in scenarios[:4])
        tweet2_text = 'How this {} {} race could unfold:\n\n{}\n\n{}'.format(course, race_time, sc_lines, url)
        queue.append({'scheduled_time':(off_dt - _timedelta(minutes=20)).isoformat(),
                      'text':tweet2_text,'type':'race_scenarios','course':course,'race_time':race_time})
    with open(TWEET_QUEUE_FILE, 'w') as f:
        _json.dump(queue, f, indent=2)
    print('  Tweet queue updated \u2014 {} tweets queued'.format(len(queue)))
    # PATCH 5b: email tweet copy after each meeting
    _send_tweet_email(queue, course, race_date_str)


def publish_meeting(meeting_results, course, race_date_str, going_report=None, generate_narratives=True):
    if not meeting_results: print('  skip {} — no results'.format(course)); return None
    if going_report is None:
        going_report = '\n'.join(m.get('going_report','') for m,_ in meeting_results if m.get('going_report')) or 'Going report not available.'
    slug = (course.lower().replace(' ','_').replace('(','').replace(')','').replace('/','_').strip('_')
            + '_' + race_date_str.lower().replace(' ','-'))
    print('\n  {} \u2014 {}'.format(course, race_date_str))
    print('  {} races | slug: {}\n'.format(len(meeting_results), slug))
    races_data = []
    for i, (meta, ev) in enumerate(meeting_results):
        is_flat = _is_flat(meta); dist_f = meta.get('dist_f',10.0)
        horses_list = meta.get('horses',[]); today_going = meta.get('going','GOOD')
        course_name = meta.get('course','')
        is_aw = any(x in course_name.upper() for x in ['(AW)','ALL WEATHER','ALL-WEATHER','DUNDALK','LINGFIELD','KEMPTON','WOLVERHAMPTON','CHELMSFORD','NEWCASTLE'])
        style_summary = _build_style_summary(horses_list, today_going, dist_f, is_aw=is_aw) if horses_list else {}
        runners = _build_runner_data(meta, ev, style_summary)
        unknown_count = sum(1 for r in runners if r.get('style_code') == 'U')
        unraced_pct = unknown_count / len(runners) if runners else 0
        if unraced_pct >= 0.40:
            print('  [{}/{}] {} {} — SKIPPED ({}% unknown style)'.format(
                i+1, len(meeting_results), meta['time'], meta['name'][:42], round(unraced_pct*100)))
            races_data.append({
                'id':i+1,'time':meta['time'],'name':meta['name'],'grade':meta['grade'],
                'dist':meta['dist'],'dist_f':dist_f,
                'going':meta.get('going','').replace('_',' ').title(),
                'runners':ev.field_size,'free':False,'type':'flat' if is_flat else 'jumps',
                'pace':'UNKNOWN','paceConf':0,'leads':[],'prominent':[],'midfield':[],'holdup':[],
                'drawBias':None,'runners_data':runners,'paceDynamic':'','scenarios':[],'watchPoints':[],'skipped':True})
            continue
        leads=[]; prominent=[]; midfield=[]; holdup=[]
        pr = meta.get('pace_result')
        if pr and hasattr(pr,'profiles'):
            for p in pr.profiles:
                rs = str(getattr(p,'assigned_style','UNKNOWN')); n = p.horse_name
                if 'FRONT' in rs: leads.append(n)
                elif 'PROMINENT' in rs: prominent.append(n)
                elif 'MIDFIELD' in rs: midfield.append(n)
                elif 'HOLD' in rs: holdup.append(n)
        pd_str, pace_info = _build_pace_dynamic(runners, style_summary, ev, meta)
        leads = pace_info['leader_names']; prominent = pace_info['prominent_names']
        midfield = pace_info['midfield_names']; holdup = pace_info['holdup_names']
        pace_dynamic_str = ''; scenarios = []; watch_points = []
        if generate_narratives:
            print('  [{}/{}] {} {}...'.format(i+1, len(meeting_results), meta['time'], meta['name'][:42]))
            try:
                prompt = _narrative_prompt(meta, ev, runners, going_report, is_flat,
                                           style_summary, pd_str, pace_info, horses_list=horses_list)
                raw = _call_claude(prompt); time.sleep(8)
                pace_dynamic_str, scenarios, notes_map, watch_points = _parse_narrative(raw, runners)
                for r in runners: r['note'] = notes_map.get(r['name'],'')
                print('         done ({} scenarios, {} watch points)'.format(len(scenarios), len(watch_points)))
            except Exception as e:
                import traceback; traceback.print_exc()
                print('         Claude error: {}'.format(e))
        else:
            print('  [{}/{}] {} {} (no narrative)'.format(i+1, len(meeting_results), meta['time'], meta['name'][:42]))
        races_data.append({
            'id':i+1,'time':meta['time'],'name':meta['name'],'grade':meta['grade'],
            'dist':meta['dist'],'dist_f':dist_f,
            'going':meta.get('going','').replace('_',' ').title(),
            'runners':ev.field_size,'free':False,'type':'flat' if is_flat else 'jumps',
            'pace':ev.pace_scenario or 'UNKNOWN','paceConf':round((ev.pace_prob or 0.5)*100),
            'leads':leads,'prominent':prominent,'midfield':midfield,'holdup':holdup,
            'drawBias':_draw_bias_summary(meta.get('course',''),dist_f) if is_flat else None,
            'runners_data':runners,'paceDynamic':pace_dynamic_str,'scenarios':scenarios,
            'watchPoints':watch_points,'skipped':False})
      
    conflicts = _consistency_check(races_data)
    if conflicts:
        print('\n  \u26a0 STYLE CONFLICTS DETECTED ({}):'.format(len(conflicts)))
        for c in conflicts:
            print('    {} {} \u2014 engine says {}, note reads {}: "{}"'.format(
                c['race'], c['horse'], c['engine'], c['note_says'], c['note'][:80]))
        flagged_times = {c['race'] for c in conflicts}
        for race in races_data:
            if race.get('time') in flagged_times:
                race['review_flag'] = True
                race['review_reason'] = 'style/note conflict — check before trusting notes'

    for race in races_data:
        if not race.get('skipped'): race['free'] = True; break
    json_str = json.dumps({'course':course,'date':race_date_str,'slug':slug,'races':races_data},
                          ensure_ascii=False, indent=2)
    print('\n  Pushing public/data/{}.json...'.format(slug))
    if not _gh_push('public/data/{}.json'.format(slug), json_str, 'data: {} {}'.format(course, race_date_str)):
        print('  JSON push FAILED'); return None
    print('  JSON pushed OK')
    html_path = 'public/meetings/{}.html'.format(slug)
    stale_url = 'https://api.github.com/repos/{}/contents/{}'.format(GITHUB_REPO, html_path)
    hdr = {'Authorization':'token {}'.format(GITHUB_TOKEN),'Accept':'application/vnd.github.v3+json'}
    stale_r = requests.get(stale_url, headers=hdr)
    if stale_r.status_code == 200:
        requests.delete(stale_url, headers=hdr, json={'message':'cleanup: {}'.format(html_path),'sha':stale_r.json().get('sha')})
        print('  Stale HTML removed')
    going_display = next((r['going'] for r in races_data if not r.get('skipped')), '')
    ok = _update_archive(slug, course, race_date_str, going_display, len(races_data))
    print('  Archive {}'.format('updated OK' if ok else 'update FAILED'))
    print(('\n'+'='*60+'\n  LIVE: pacemap.co.uk/meetings/{}\n  (Vercel ~60s)\n'+'='*60+'\n').format(slug))
    first_valid = next((r for r in races_data if not r.get('skipped')), None)
    if first_valid:
        time_str = first_valid['time']; dist = first_valid['dist']
        going = first_valid['going']; n_runners = first_valid['runners']
        dynamic = first_valid.get('paceDynamic',''); scens = first_valid.get('scenarios',[])
        url = 'pacemap.co.uk/meetings/{}'.format(slug)
        short_dynamic = (dynamic[:180] + '...') if len(dynamic) > 180 else dynamic
        sc_lines = '\n'.join('{}) {} ({}%) — {} benefit'.format(
            s['label'],s['title'],s['prob'],', '.join(s['winners'][:2])) for s in scens[:4])
        print('\n' + '\u2501'*50)
        print('📋 TWEET COPY — {} {}'.format(course, time_str))
        print('\u2501'*50)
        print('TWEET 1:')
        print('📍 {} {} \u00b7 {} \u00b7 {} \u00b7 {} runners\n\n{}\n\n🔍 {}'.format(
            course, time_str, dist, going, n_runners, short_dynamic, url))
        print('\nTWEET 2:')
        print('How this race could unfold:\n\n{}'.format(sc_lines))
        print('\u2501'*50 + '\n')
    _write_tweet_queue(slug, course, race_date_str, races_data)
    return slug

def publish_all(all_results=None, race_date_str=None, going_report=None, generate_narratives=True):
    if all_results is None: all_results = ALL_RESULTS
    if race_date_str is None:
        try:
            import pytz; uk_now = datetime.datetime.now(pytz.timezone('Europe/London'))
        except ImportError:
            uk_now = datetime.datetime.utcnow() + datetime.timedelta(hours=1)
        if uk_now.hour >= 18:
            target = uk_now.date() + datetime.timedelta(days=1)
            print('  Running after 6pm UK time — publishing for tomorrow ({})'.format(target.strftime('%-d %B %Y')))
        else:
            target = uk_now.date()
        race_date_str = target.strftime('%-d %B %Y')
    if _os.path.exists(TWEET_QUEUE_FILE): _os.remove(TWEET_QUEUE_FILE)
    slugs = []
    for course, results in all_results.items():
        slug = publish_meeting(results, course, race_date_str, going_report=going_report, generate_narratives=generate_narratives)
        if slug: slugs.append(slug)
        time.sleep(1)
    print('\nPublished {} meeting(s) for {}'.format(len(slugs), race_date_str))
    return slugs

_active = {k:v for k,v in ALL_RESULTS.items() if v} if 'ALL_RESULTS' in dir() else {}
if _active:
    publish_all(all_results=_active, generate_narratives=True)
else:
    print('No results found — run Cells 7 and 8 first.')
